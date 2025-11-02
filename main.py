import os
import sys
import glob
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import xlwings as xw
from typing import List, Dict, Optional
import logging
from dataclasses import dataclass

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@dataclass
class ProcessingResult:
    """Data class for processing results"""
    success: bool
    file_name: str
    records_found: int
    data: Optional[List[Dict]] = None
    error_message: Optional[str] = None


class MacroManager:
    """Manage external VBA macro files"""
    
    def __init__(self, macro_dir: str = "macros"):
        self.macro_dir = Path(macro_dir).resolve()
        self.macro_files = {
            'module2': 'Module2.bas',
            'module3': 'Module3.bas',
            'module5': 'Module5.bas'
        }
        
    def inject_macros(self, workbook) -> bool:
        """Inject external VBA macros into workbook"""
        try:
            for module_name, file_name in self.macro_files.items():
                macro_path = self.macro_dir / file_name
                
                if not macro_path.exists():
                    logger.warning(f"Macro file not found: {macro_path}")
                    continue
                
                # Read macro content
                with open(macro_path, 'r', encoding='utf-8') as f:
                    macro_code = f.read()
                
                # Check if module exists, remove if it does
                try:
                    workbook.api.VBProject.VBComponents.Remove(
                        workbook.api.VBProject.VBComponents(module_name)
                    )
                except:
                    pass
                
                # Import new module with absolute path
                vb_module = workbook.api.VBProject.VBComponents.Import(str(macro_path))
                logger.info(f"Injected macro: {module_name}")
                
            return True
            
        except Exception as e:
            logger.error(f"Failed to inject macros: {e}")
            return False
    
    def verify_macros_exist(self, workbook) -> Dict[str, bool]:
        """Verify which macros exist in workbook"""
        macro_status = {}
        
        required_macros = [
            'RemoveDuplicatesFromTables',
            'SummarizeAchievement',
            'Save_Range_As_PNG_And_Export_XLSX',
            'ProcessZeroValuesSeparate'
        ]
        
        for macro_name in required_macros:
            try:
                workbook.macro(macro_name)
                macro_status[macro_name] = True
            except:
                macro_status[macro_name] = False
                
        return macro_status


class XLSBProcessor:
    """Core processor for XLSB files"""
    
    def __init__(self, use_external_macros: bool = True):
        self.use_external_macros = use_external_macros
        self.macro_manager = MacroManager() if use_external_macros else None
        
    def process_file(self, file_path: str, log_callback=None) -> ProcessingResult:
        """Process single XLSB file"""
        app = None
        wb = None
        output_data = []
        
        try:
            file_name = Path(file_path).name
            self._log(f"Opening: {file_name}", log_callback)
            
            # Open Excel with hidden app
            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False
            app.screen_updating = False
            
            wb = app.books.open(file_path)
            
            # Inject external macros if enabled
            if self.use_external_macros:
                self._log("Injecting external macros...", log_callback)
                if not self.macro_manager.inject_macros(wb):
                    return ProcessingResult(
                        success=False,
                        file_name=file_name,
                        records_found=0,
                        data=None,
                        error_message="Failed to inject macros"
                    )
            
            # Verify macros
            macro_status = self.macro_manager.verify_macros_exist(wb)
            missing_macros = [k for k, v in macro_status.items() if not v]
            
            if missing_macros:
                self._log(f"Missing macros: {missing_macros}", log_callback)
            
            # Get worksheets
            ws_final = wb.sheets["Final Summary"]
            ws_achievement = wb.sheets["SSH Achivement"]
            
            # Get date range
            min_date, max_date = self._get_date_range(ws_achievement)
            self._log(f"Date range: {min_date} to {max_date}", log_callback)
            ws_achievement.range("AB23").value = min_date
            ws_achievement.range("AB24").value = max_date
            
            # Step 1: Remove duplicates (silent execution)
            self._log("Step 1: Removing duplicates...", log_callback)
            self._run_macro_silent(wb, "RemoveDuplicatesFromTables", log_callback)
            
            # Step 3: Get candidate Q sites (blank or = Max Date)
            candidate_sites = self._get_candidate_q_sites(ws_final, max_date, log_callback)
            
            if not candidate_sites:
                self._log("No candidate Q values found. All sites processed.", log_callback)
                wb.save()
                return ProcessingResult(
                    success=True,
                    file_name=file_name,
                    records_found=0,
                    data=[]
                )
            
            # Step 4-11: Process clusters and sites
            grouped_sites = self._group_by_cluster(candidate_sites)
            
            for cluster_name, sites in grouped_sites.items():
                self._log(f"Processing cluster: {cluster_name} ({len(sites)} sites)", log_callback)
                ws_achievement.range("C2").value = cluster_name
                
                for site_info in sites:
                    site_output = self._process_site(
                        wb, ws_final, ws_achievement, 
                        site_info, max_date, min_date, log_callback
                    )
                    output_data.extend(site_output)
            
            # Save workbook
            wb.save()
            self._log("Workbook saved successfully", log_callback)
            
            return ProcessingResult(
                success=True,
                file_name=file_name,
                records_found=len(output_data),
                data=output_data
            )
            
        except Exception as e:
            logger.exception(f"Error processing {file_path}")
            return ProcessingResult(
                success=False,
                file_name=Path(file_path).name,
                records_found=0,
                data=None,
                error_message=str(e)
            )
            
        finally:
            if wb:
                wb.close()
            if app:
                app.quit()
    
    def _run_macro_silent(self, wb, macro_name: str, log_callback=None):
        """Run macro silently without popups"""
        try:
            # Disable alerts
            wb.app.display_alerts = False
            wb.app.screen_updating = False
            
            # Run macro
            macro = wb.macro(macro_name)
            macro()
            
            self._log(f"✓ {macro_name} completed", log_callback)
            
        except Exception as e:
            self._log(f"⚠ {macro_name} error: {str(e)}", log_callback)
        finally:
            wb.app.display_alerts = False
    
    def _get_date_range(self, ws_achievement):
        """Get min and max dates from dropdown"""
        try:
            validation = ws_achievement.range("AB24").api.Validation
            if validation.Type == 3:
                formula = validation.Formula1
                if "=" in formula:
                    range_address = formula.split("=")[1]
                    dropdown_values = ws_achievement.range(range_address).value
                    valid_dates = []
                    if isinstance(dropdown_values, list):
                        valid_dates = [d for d in dropdown_values if d is not None and d != ""]
                    else:
                        if dropdown_values is not None and dropdown_values != "":
                            valid_dates = [dropdown_values]
                    if valid_dates:
                        return min(valid_dates), max(valid_dates)
            min_d = ws_achievement.range("AB23").value
            max_d = ws_achievement.range("AB24").value
            return min_d or max_d, max_d
        except:
            return ws_achievement.range("AB23").value, ws_achievement.range("AB24").value
    
    def _get_candidate_q_sites(self, ws_final, max_date, log_callback=None) -> List[Dict]:
        """Get sites with blank Q column or Q = max_date"""
        candidate_sites = []
        last_row = ws_final.range('B' + str(ws_final.cells.last_cell.row)).end('up').row
        seen = set()
        
        for row in range(7, last_row + 1):
            q_val = ws_final.range(f"Q{row}").value
            a_val = ws_final.range(f"A{row}").value
            b_val = ws_final.range(f"B{row}").value
            
            is_candidate = False
            if not q_val or str(q_val).strip() == "":
                is_candidate = True
            elif self._compare_dates(q_val, max_date):
                is_candidate = True
            
            if is_candidate and a_val and b_val:
                combo = f"{a_val}|{b_val}"
                if combo not in seen:
                    seen.add(combo)
                    candidate_sites.append({
                        'row': row,
                        'cluster': a_val,
                        'site': b_val
                    })
        
        return candidate_sites
    
    def _group_by_cluster(self, sites: List[Dict]) -> Dict[str, List[Dict]]:
        """Group sites by cluster"""
        grouped = {}
        for site in sites:
            cluster = site['cluster']
            if cluster not in grouped:
                grouped[cluster] = []
            grouped[cluster].append(site)
        return grouped
    
    def _process_site(self, wb, ws_final, ws_achievement, 
                     site_info, max_date, min_date, log_callback) -> List[Dict]:
        """Process individual site"""
        output = []
        
        try:
            # Set site
            ws_achievement.range("C6").value = site_info['site']
            
            # Get integration date
            integration_date = ws_final.range(f"P{site_info['row']}").value
            if not integration_date:
                return output
            
            # Set integration date for calculations
            ws_achievement.range("AA6").value = integration_date
            
            # Check current Q value
            q_value = ws_final.range(f"Q{site_info['row']}").value
            run_summarize = not q_value or str(q_value).strip() == ""
            
            if run_summarize:
                # Run SummarizeAchievement to fill Q/R if needed
                self._run_macro_silent(wb, "SummarizeAchievement", log_callback)
                q_value = ws_final.range(f"Q{site_info['row']}").value
            
            if not q_value or not self._compare_dates(q_value, max_date):
                return output
            
            # Formulas for baselines
            formulas = [
                "=AB14-((ROUNDUP((AB14-(AA6-1))/7,0))*7)",
                "=AB14-((ROUNDUP((AB14-(AA6-7))/7,0))*7)",
                "=AB14-((ROUNDUP((AB14-(AA6-14))/7,0))*7)"
            ]
            
            # Set max date
            ws_achievement.range("AB14").value = max_date
            wb.app.calculate()
            
            for case in range(1, 4):
                formula = formulas[case - 1]
                ws_achievement.range("AB10").formula = formula
                wb.app.calculate()
                
                baseline_date_raw = ws_achievement.range("AB10").value
                if not baseline_date_raw:
                    continue
                
                if self._check_pass_status(ws_final, site_info['row'], case):
                    self._log(f"  ✓ PASS for Case {case}", log_callback)
                    
                    # Export XLSX
                    self._run_macro_silent(wb, "Save_Range_As_PNG_And_Export_XLSX", log_callback)
                    
                    # Get remark for this case
                    remark_str = self._get_remark_for_case(ws_achievement)
                    
                    # Extract data for this case
                    data = self._extract_site_data_for_case(
                        ws_final, site_info['row'], max_date, integration_date, 
                        q_value, baseline_date_raw, remark_str
                    )
                    if data:
                        output.append(data)
        
        except Exception as e:
            self._log(f"Error processing site: {e}", log_callback)
        
        return output
    
    def _check_pass_status(self, ws_final, row, case) -> bool:
        """Check if site has pass status for specific case"""
        col_map = {1: 'M', 2: 'N', 3: 'O'}
        col = col_map.get(case, 'O')
        col_val = ws_final.range(f"{col}{row}").value
        return col_val == "Pass"
    
    def _get_remark_for_case(self, ws_achievement) -> str:
        """Get remark text by checking zero values"""
        kpi_list = []
        band_cells = {
            'P': 'N8',
            'S': 'Q8',
            'V': 'T8',
            'Y': 'W8'
        }
        data_cols = ['P', 'S', 'V', 'Y']
        
        for i in range(10, 29):
            kpi_name = ws_achievement.range(f"M{i}").value
            if not kpi_name:
                continue
            
            for data_col in data_cols:
                val = ws_achievement.range(f"{data_col}{i}").value
                if self._is_zero_value(val):
                    band_cell = band_cells[data_col]
                    band = ws_achievement.range(band_cell).value or ""
                    kpi_combined = f"{kpi_name} {band}".strip()
                    if kpi_combined:
                        kpi_list.append(kpi_combined)
        
        if kpi_list:
            return " | ".join(kpi_list)
        return ""
    
    def _is_zero_value(self, val) -> bool:
        """Check if value is zero or <=1%"""
        if val is None or val == "":
            return False
        if isinstance(val, (int, float)):
            return abs(val) <= 1
        if isinstance(val, str):
            clean = val.strip().rstrip('%')
            try:
                return abs(float(clean)) <= 1
            except ValueError:
                return False
        return False
    
    def _compare_dates(self, date1, date2) -> bool:
        """Compare two dates"""
        try:
            if isinstance(date1, datetime) and isinstance(date2, datetime):
                return date1.date() == date2.date()
            d1 = pd.to_datetime(date1).date() if date1 else None
            d2 = pd.to_datetime(date2).date() if date2 else None
            return d1 == d2
        except:
            return False
    
    def _extract_site_data_for_case(self, ws_final, row, max_date, integration_date_raw, 
                                   pass_date_raw, baseline_date_raw, remark_str) -> Optional[Dict]:
        """Extract site data for specific baseline case"""
        try:
            cluster = ws_final.range(f"A{row}").value
            tower_id = ws_final.range(f"B{row}").value
            
            # Format dates
            if isinstance(integration_date_raw, datetime):
                integration_date = integration_date_raw.strftime("%m/%d/%Y")
            else:
                integration_date = ""
            
            if isinstance(pass_date_raw, datetime):
                pass_date = pass_date_raw.strftime("%m/%d/%Y")
            else:
                pass_date = ""
            
            if isinstance(baseline_date_raw, datetime):
                baseline_date = baseline_date_raw.strftime("%m/%d/%Y")
            else:
                baseline_date = ""
            
            # Calculate aging
            aging = ""
            if integration_date_raw and max_date:
                try:
                    int_date = pd.to_datetime(integration_date_raw)
                    mx_date = pd.to_datetime(max_date)
                    aging = (mx_date - int_date).days
                except:
                    pass
            
            # Status and flag
            status = "PASS"
            flag = "PASS WITH REMARK" if remark_str else "PASS"
            
            return {
                "CLUSTER": cluster or "",
                "TOWERID": tower_id or "",
                "Integration Date": integration_date or "",
                "AGING (Days)": aging,
                "Pass Date": pass_date or "",
                "Baseline Date": baseline_date or "",
                "FLAG": flag,
                "Status": status,
                "Remark": remark_str
            }
            
        except Exception as e:
            logger.error(f"Error extracting data: {e}")
            return None
    
    def _log(self, message: str, callback=None):
        """Log message"""
        logger.info(message)
        if callback:
            callback(message)


class ModernGUI:
    """Modern GUI for XLSB Processor"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("SSH Achievement Multi-Site Processor")
        self.root.geometry("900x700")
        
        self.processor = XLSBProcessor(use_external_macros=True)
        self.output_data = []
        
        self._setup_styles()
        self._create_widgets()
        
    def _setup_styles(self):
        """Setup modern styles"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors
        style.configure('Title.TLabel', 
                       font=('Segoe UI', 18, 'bold'),
                       foreground='#2C3E50')
        
        style.configure('Header.TLabel',
                       font=('Segoe UI', 10, 'bold'),
                       foreground='#34495E')
        
        style.configure('Action.TButton',
                       font=('Segoe UI', 10),
                       padding=10)
        
    def _create_widgets(self):
        """Create UI widgets"""
        # Main container
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title
        title = ttk.Label(main_frame, 
                         text="SSH Achievement Multi-Site Processor",
                         style='Title.TLabel')
        title.grid(row=0, column=0, columnspan=3, pady=(0, 30))
        
        # Folder selection
        folder_frame = ttk.LabelFrame(main_frame, text="Folder Selection", padding="15")
        folder_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 20))
        
        self.folder_var = tk.StringVar()
        folder_entry = ttk.Entry(folder_frame, textvariable=self.folder_var, width=70)
        folder_entry.grid(row=0, column=0, padx=(0, 10))
        
        browse_btn = ttk.Button(folder_frame, text="Browse", command=self._browse_folder)
        browse_btn.grid(row=0, column=1)
        
        # Process button
        self.process_btn = ttk.Button(main_frame,
                                      text="🚀 Start Processing",
                                      command=self._start_processing,
                                      style='Action.TButton',
                                      state='disabled')
        self.process_btn.grid(row=2, column=0, columnspan=3, pady=20)
        
        # Progress
        progress_frame = ttk.Frame(main_frame)
        progress_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress.pack(fill=tk.X)
        
        self.status_label = ttk.Label(progress_frame, text="Ready")
        self.status_label.pack(pady=(5, 0))
        
        # Log area
        log_frame = ttk.LabelFrame(main_frame, text="Processing Log", padding="10")
        log_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        
        # Text widget with scrollbar
        text_frame = ttk.Frame(log_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(text_frame, height=20, wrap=tk.WORD,
                               font=('Consolas', 9))
        scrollbar = ttk.Scrollbar(text_frame, command=self.log_text.yview)
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(4, weight=1)
        
    def _browse_folder(self):
        """Browse for folder"""
        folder = filedialog.askdirectory(title="Select Folder with XLSB Files")
        if folder:
            self.folder_var.set(folder)
            xlsb_files = glob.glob(os.path.join(folder, "*.xlsb"))
            
            if xlsb_files:
                self.process_btn.config(state='normal')
                self.status_label.config(text=f"Found {len(xlsb_files)} XLSB files")
                self._log(f"Folder selected: {folder}")
                self._log(f"Found {len(xlsb_files)} XLSB files")
            else:
                self.process_btn.config(state='disabled')
                self.status_label.config(text="No XLSB files found")
                self._log("No XLSB files found in folder")
    
    def _start_processing(self):
        """Start processing files"""
        folder = self.folder_var.get()
        if not folder or not os.path.exists(folder):
            messagebox.showerror("Error", "Invalid folder!")
            return
        
        xlsb_files = glob.glob(os.path.join(folder, "*.xlsb"))
        if not xlsb_files:
            messagebox.showwarning("Warning", "No XLSB files found!")
            return
        
        self.output_data = []
        self.process_btn.config(state='disabled')
        self.progress['maximum'] = len(xlsb_files)
        self.progress['value'] = 0
        
        results = []
        
        for i, file_path in enumerate(xlsb_files):
            self.status_label.config(text=f"Processing: {Path(file_path).name}")
            self._log(f"\n{'='*70}")
            self._log(f"Processing: {Path(file_path).name}")
            
            result = self.processor.process_file(file_path, self._log)
            results.append(result)
            
            if result.success:
                if result.data:
                    self.output_data.extend(result.data)
                self._log(f"✓ Success - {result.records_found} records")
            else:
                self._log(f"✗ Failed - {result.error_message}")
            
            self.progress['value'] = i + 1
            self.root.update_idletasks()
        
        # Summary
        success_count = sum(1 for r in results if r.success)
        error_count = len(results) - success_count
        
        self.status_label.config(
            text=f"Complete! Success: {success_count}, Failed: {error_count}"
        )
        self._log(f"\n{'='*70}")
        self._log(f"SUMMARY: {success_count} success, {error_count} failed")
        
        # Save output
        if self.output_data:
            self._save_output(folder)
        
        self.process_btn.config(state='normal')
        
        if error_count == 0:
            messagebox.showinfo("Success", 
                              f"All {success_count} files processed successfully!")
        else:
            messagebox.showwarning("Completed",
                                  f"Processing complete!\nSuccess: {success_count}\nFailed: {error_count}")
    
    def _save_output(self, folder):
        """Save output Excel"""
        try:
            df = pd.DataFrame(self.output_data)
            if not df.empty:
                df.insert(0, 'No', range(1, len(df) + 1))
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"SUMMARY_SSH_CLUSTER_{timestamp}.xlsx"
                output_path = os.path.join(folder, output_file)
                
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="Summary")
                    
                    from openpyxl.styles import PatternFill, Font, Alignment
                    
                    ws = writer.sheets["Summary"]
                    
                    # Header styling
                    for cell in ws[1]:
                        cell.font = Font(bold=True, size=11)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Status coloring
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    status_col = df.columns.get_loc("Status") + 1  # 1-based column index
                    for row in range(2, len(df) + 2):
                        status_cell = ws.cell(row=row, column=status_col)
                        if status_cell.value == "PASS":
                            status_cell.fill = green_fill
                        elif status_cell.value == "FAIL":
                            status_cell.fill = red_fill
                    
                    # Auto-adjust columns
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                
                self._log(f"\n✓ Output saved: {output_file}")
            else:
                self._log("\nNo data to save.")
                
        except Exception as e:
            self._log(f"✗ Error saving output: {e}")
    
    def _log(self, message):
        """Add log message"""
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()


def main():
    root = tk.Tk()
    app = ModernGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()