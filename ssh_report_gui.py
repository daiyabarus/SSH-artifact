import os
import re
import sqlite3
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Tuple, Dict
from io import BytesIO
import warnings

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from copy import copy

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

import subprocess
import time

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


@dataclass
class KPIConfig:
    """Configuration for a single KPI metric"""

    column_name: str
    source_df_name: str
    band_column: Optional[str]
    date_column: str
    title: str
    unit: str
    ylim_min: Optional[float] = None
    ylim_max: Optional[float] = None


@dataclass
class BandConfig:
    """Band-related configuration"""

    colors: Dict[str, str] = field(
        default_factory=lambda: {
            "850": "#52eb0c",
            "1800": "#080cec",
            "2100": "#ef17e8",
            "2300F1": "#F39C12",
            "2300F2": "#9B59B6",
        }
    )

    columns: Dict[str, str] = field(
        default_factory=lambda: {
            "M": "LowBand",
            "P": "MidBand_18",
            "S": "MidBand_21",
            "V": "HighBand_23",
        }
    )

    values: Dict[str, List[int]] = field(
        default_factory=lambda: {
            "LowBand": [850],
            "MidBand_18": [1800],
            "MidBand_21": [2100],
            "HighBand_23": [2300],
        }
    )


class SSHReportGenerator:
    def __init__(
        self, input_folder: str, template_path: str, output_folder: str, db_path: str
    ):
        self.input_folder = Path(input_folder).resolve()
        self.template_path = Path(template_path).resolve()
        self.output_folder = Path(output_folder).resolve()
        self.db_path = Path(db_path).resolve()
        self.output_folder.mkdir(exist_ok=True, parents=True)

        # Color palette for general use
        self.color_palette = [
            "#080cec",
            "#ef17e8",
            "#52eb0c",
            "#f59e0b",
            "#10b981",
            "#06b6d4",
            "#ef4444",
            "#a855f7",
            "#14b8a6",
            "#f97316",
        ]

        self.band_config = BandConfig()
        self._init_kpi_configs()

    def _init_kpi_configs(self):
        """Initialize KPI configurations"""
        self.kpi_mapping = {
            11: "Session Setup Success Rate (%)",
            12: "Volte CSSR (%)",
            13: "Erab Drop (%)",
            14: "HOSR (Intra + Inter, LTE to LTE) (%)",
            15: "CQI",
            16: "QPSK Distribution (%)",
            17: "Spectrum Efficiency (bps/Hz)",
            18: "Overall Packet Loss Rate (%)",
            19: "Overall Latency",
            20: "DL Cell Throughput",
            21: "UL Cell Throughput",
            22: "Rank2 (%)",
            23: "UL RSSI (dBm)",
            24: "Last TTI (%)",
        }

        self.kpi_configs = {
            15: KPIConfig(
                "cqi",
                "bh",
                "newbh_cell_fdd_band",
                "newbh_date",
                "Average CQI",
                "CQI",
                0,
                15,
            ),
            16: KPIConfig(
                "qpsk",
                "bh",
                "newbh_cell_fdd_band",
                "newbh_date",
                "QPSK Rate",
                "%",
                0,
                100,
            ),
            17: KPIConfig(
                "spectral_eff",
                "wd",
                "newwd_cell_fdd_band",
                "newwd_date",
                "Spectral Efficiency",
                "bps/Hz",
                0,
                None,
            ),
            18: KPIConfig(
                "packet_loss",
                "kqi",
                None,
                "newkqi_date",
                "Overall Packet Loss Rate",
                "%",
                0,
                None,
            ),
            19: KPIConfig(
                "latency", "kqi", None, "newkqi_date", "Overall Latency", "ms", 0, None
            ),
            20: KPIConfig(
                "dl_throughput",
                "bh",
                "newbh_cell_fdd_band",
                "newbh_date",
                "DL Cell Throughput",
                "Mbps",
                0,
                None,
            ),
            21: KPIConfig(
                "ul_throughput",
                "bh",
                "newbh_cell_fdd_band",
                "newbh_date",
                "UL Cell Throughput",
                "Mbps",
                0,
                None,
            ),
            22: KPIConfig(
                "rank2", "bh", "newbh_cell_fdd_band", "newbh_date", "Rank2", "%", 0, 100
            ),
            23: KPIConfig(
                "ul_rssi",
                "wd",
                "newwd_cell_fdd_band",
                "newwd_date",
                "UL RSSI",
                "dBm",
                None,
                None,
            ),
            24: KPIConfig(
                "last_tti",
                "bh",
                "newbh_cell_fdd_band",
                "newbh_date",
                "Last TTI",
                "%",
                0,
                100,
            ),
        }

    @contextmanager
    def db_connection(self):
        """Context manager for database connections"""
        if not self.db_path.exists():
            raise FileNotFoundError(f"Database not found: {self.db_path}")
        conn = sqlite3.connect(self.db_path)
        try:
            yield conn
        finally:
            conn.close()

    @staticmethod
    def parse_date_flexible(date_str) -> Optional[datetime]:
        """Parse date from various formats"""
        if pd.isna(date_str):
            return None

        date_formats = [
            "%m/%d/%Y",
            "%Y-%m-%d",
            "%m/%d/%y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%d-%m-%Y",
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(str(date_str).strip(), fmt)
            except ValueError:
                continue

        print(f"  ⚠ Could not parse date: {date_str}")
        return None

    def query_data(
        self, tower_id: str
    ) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """Query all required data from database"""
        with self.db_connection() as conn:
            ta_df = self._query_ta_data(conn, tower_id)
            wd_df = self._query_wd_data(conn, tower_id)
            bh_df = self._query_bh_data(conn, tower_id)
            kqi_df = self._query_kqi_data(conn, tower_id)

        return ta_df, wd_df, bh_df, kqi_df

    def _query_ta_data(self, conn, tower_id: str) -> pd.DataFrame:
        """Query Timing Advance data"""
        query = """
        SELECT * FROM tbl_newta 
        WHERE newta_managed_element = ?
        ORDER BY newta_date DESC
        """
        return pd.read_sql_query(query, conn, params=(tower_id,))

    def _query_wd_data(self, conn, tower_id: str) -> pd.DataFrame:
        """Query WD (Daily) data"""
        query = """
        SELECT newwd_date, newwd_moentity, newwd_cell_fdd_band,
               newwd_enodeb_fdd_msc, newwd_operator, newwd_cell_fdd_system,
               newwd_cell_fdd_txrxmode, newwd_cell_fdd_vendor,
               newwd_spectral_efficiency_dl_num, newwd_spectral_efficiency_dl_den,
               newwd_ul_rssi_num_dbm, newwd_ul_rssi_denom_dbm
        FROM tbl_newwd
        WHERE newwd_enodeb_fdd_msc = ?
        ORDER BY newwd_date
        """
        df = pd.read_sql_query(query, conn, params=(tower_id,))
        df["newwd_date"] = df["newwd_date"].apply(self.parse_date_flexible)
        return df

    def _query_bh_data(self, conn, tower_id: str) -> pd.DataFrame:
        """Query BH (Busy Hour) data"""
        query = """
        SELECT newbh_date, newbh_moentity, newbh_cell_fdd_band, newbh_enodeb_fdd_msc,
               newbh_pdcp_cell_throughput_dl_num, newbh_pdcp_cell_throughput_dl_denom,
               newbh_pdcp_cell_throughput_ul_num, newbh_pdcp_cell_throughput_ul_den,
               newbh_cell_average_cqi_num, newbh_cell_average_cqi_den,
               newbh_cell_qpsk_rate_num, newbh_cell_qpsk_rate_den,
               newbh_cell_mimo_transmission_rank_eq_2_rate_num, 
               newbh_cell_mimo_transmission_rank_eq_2_rate_den,
               newbh_cell_last_tti_ratio_num, newbh_cell_last_tti_ratio_den,
               newbh_ul_rssi_num_dbm, newbh_ul_rssi_denom_dbm
        FROM tbl_newbh
        WHERE newbh_enodeb_fdd_msc = ?
        ORDER BY newbh_date
        """
        df = pd.read_sql_query(query, conn, params=(tower_id,))
        df["newbh_date"] = df["newbh_date"].apply(self.parse_date_flexible)
        return df

    def _query_kqi_data(self, conn, tower_id: str) -> pd.DataFrame:
        """Query KQI data and aggregate by date"""
        query = """
        SELECT newkqi_date, 
            newkqi_swe_l6,
            SUM(newkqi_tcp_connect_delay_ms) as newkqi_tcp_connect_delay_ms,
            SUM(newkqi_tcp_connect_rtt_count_times) as newkqi_tcp_connect_rtt_count_times,
            AVG(newkqi_server_side_uplink_tcp_packet_loss_rate) as newkqi_server_side_uplink_tcp_packet_loss_rate,
            AVG(newkqi_server_side_downlink_tcp_packet_loss_rate) as newkqi_server_side_downlink_tcp_packet_loss_rate,
            AVG(newkqi_client_side_uplink_tcp_packet_loss_rate) as newkqi_client_side_uplink_tcp_packet_loss_rate,
            AVG(newkqi_client_side_downlink_tcp_packet_loss_rate) as newkqi_client_side_downlink_tcp_packet_loss_rate
        FROM tbl_newkqi
        WHERE newkqi_swe_l6 = ?
        GROUP BY newkqi_date, newkqi_swe_l6
        ORDER BY newkqi_date
        """
        df = pd.read_sql_query(query, conn, params=(tower_id,))
        df["newkqi_date"] = df["newkqi_date"].apply(self.parse_date_flexible)
        return df

    def merge_data(
        self, ta_df: pd.DataFrame, wd_df: pd.DataFrame, bh_df: pd.DataFrame
    ) -> Tuple[pd.DataFrame, Optional[pd.DataFrame], Optional[pd.DataFrame]]:
        """Merge dataframes with timing advance info"""
        if ta_df.empty:
            return None, None, None

        wd_merged = None
        if not wd_df.empty:
            wd_merged = pd.merge(
                wd_df,
                ta_df[["newta_eutrancell", "newta_sector", "newta_sector_name"]],
                left_on="newwd_moentity",
                right_on="newta_eutrancell",
                how="left",
            )

        bh_merged = None
        if not bh_df.empty:
            bh_merged = pd.merge(
                bh_df,
                ta_df[["newta_eutrancell", "newta_sector", "newta_sector_name"]],
                left_on="newbh_moentity",
                right_on="newta_eutrancell",
                how="left",
            )

        return ta_df, wd_merged, bh_merged

    def compute_derived_metrics(
        self,
        wd_merged: Optional[pd.DataFrame],
        bh_merged: Optional[pd.DataFrame],
        kqi_df: Optional[pd.DataFrame],
    ):
        """Compute all derived metrics from raw data"""
        if bh_merged is not None and not bh_merged.empty:
            # Convert numeric columns to float
            numeric_cols = [
                "newbh_pdcp_cell_throughput_dl_num",
                "newbh_pdcp_cell_throughput_dl_denom",
                "newbh_pdcp_cell_throughput_ul_num",
                "newbh_pdcp_cell_throughput_ul_den",
                "newbh_cell_average_cqi_num",
                "newbh_cell_average_cqi_den",
                "newbh_cell_qpsk_rate_num",
                "newbh_cell_qpsk_rate_den",
                "newbh_cell_mimo_transmission_rank_eq_2_rate_num",
                "newbh_cell_mimo_transmission_rank_eq_2_rate_den",
                "newbh_cell_last_tti_ratio_num",
                "newbh_cell_last_tti_ratio_den",
            ]
            for col in numeric_cols:
                if col in bh_merged.columns:
                    bh_merged[col] = pd.to_numeric(bh_merged[col], errors="coerce")

            # Throughput
            bh_merged["dl_throughput"] = np.where(
                bh_merged["newbh_pdcp_cell_throughput_dl_denom"] > 0,
                bh_merged["newbh_pdcp_cell_throughput_dl_num"]
                / bh_merged["newbh_pdcp_cell_throughput_dl_denom"],
                np.nan,
            )
            bh_merged["ul_throughput"] = np.where(
                bh_merged["newbh_pdcp_cell_throughput_ul_den"] > 0,
                bh_merged["newbh_pdcp_cell_throughput_ul_num"]
                / bh_merged["newbh_pdcp_cell_throughput_ul_den"],
                np.nan,
            )
            # CQI
            bh_merged["cqi"] = np.where(
                bh_merged["newbh_cell_average_cqi_den"] > 0,
                bh_merged["newbh_cell_average_cqi_num"]
                / bh_merged["newbh_cell_average_cqi_den"],
                np.nan,
            )
            # QPSK
            bh_merged["qpsk"] = np.where(
                bh_merged["newbh_cell_qpsk_rate_den"] > 0,
                (
                    bh_merged["newbh_cell_qpsk_rate_num"]
                    / bh_merged["newbh_cell_qpsk_rate_den"]
                )
                * 100,
                np.nan,
            )
            # Rank2
            bh_merged["rank2"] = np.where(
                bh_merged["newbh_cell_mimo_transmission_rank_eq_2_rate_den"] > 0,
                (
                    bh_merged["newbh_cell_mimo_transmission_rank_eq_2_rate_num"]
                    / bh_merged["newbh_cell_mimo_transmission_rank_eq_2_rate_den"]
                )
                * 100,
                np.nan,
            )
            # Last TTI
            bh_merged["last_tti"] = np.where(
                bh_merged["newbh_cell_last_tti_ratio_den"] > 0,
                (
                    bh_merged["newbh_cell_last_tti_ratio_num"]
                    / bh_merged["newbh_cell_last_tti_ratio_den"]
                )
                * 100,
                np.nan,
            )

        if wd_merged is not None and not wd_merged.empty:
            # Convert numeric columns to float
            numeric_cols = [
                "newwd_spectral_efficiency_dl_num",
                "newwd_spectral_efficiency_dl_den",
                "newwd_ul_rssi_num_dbm",
                "newwd_ul_rssi_denom_dbm",
            ]
            for col in numeric_cols:
                if col in wd_merged.columns:
                    wd_merged[col] = pd.to_numeric(wd_merged[col], errors="coerce")

            # Spectral Efficiency
            wd_merged["spectral_eff"] = np.where(
                wd_merged["newwd_spectral_efficiency_dl_den"] > 0,
                wd_merged["newwd_spectral_efficiency_dl_num"]
                / wd_merged["newwd_spectral_efficiency_dl_den"],
                np.nan,
            )
            # UL RSSI
            wd_merged["ul_rssi"] = np.where(
                wd_merged["newwd_ul_rssi_denom_dbm"] > 0,
                wd_merged["newwd_ul_rssi_num_dbm"]
                / wd_merged["newwd_ul_rssi_denom_dbm"],
                np.nan,
            )

        if kqi_df is not None and not kqi_df.empty:
            # Convert numeric columns to float
            numeric_cols = [
                "newkqi_tcp_connect_delay_ms",
                "newkqi_tcp_connect_rtt_count_times",
                "newkqi_server_side_uplink_tcp_packet_loss_rate",
                "newkqi_server_side_downlink_tcp_packet_loss_rate",
            ]
            for col in numeric_cols:
                if col in kqi_df.columns:
                    kqi_df[col] = pd.to_numeric(kqi_df[col], errors="coerce")

            # Latency - already aggregated by SQL SUM
            kqi_df["latency"] = np.where(
                kqi_df["newkqi_tcp_connect_rtt_count_times"] > 0,
                kqi_df["newkqi_tcp_connect_delay_ms"]
                / kqi_df["newkqi_tcp_connect_rtt_count_times"],
                np.nan,
            )
            # Packet Loss - already aggregated by SQL AVG
            kqi_df["packet_loss"] = (
                (
                    kqi_df["newkqi_server_side_uplink_tcp_packet_loss_rate"]
                    + kqi_df["newkqi_server_side_downlink_tcp_packet_loss_rate"]
                )
                / 2
                * 100
            )

    @staticmethod
    def is_cell_zero_value(cell) -> bool:
        """Check if a cell value is effectively zero"""
        if cell.value is None:
            return False

        value = cell.value

        # Numeric check
        if isinstance(value, (int, float)):
            try:
                return abs(float(value)) < 1e-10
            except (ValueError, TypeError):
                return False

        # String check
        if isinstance(value, str):
            str_val = value.strip()
            if not str_val:
                return False

            simple_zeros = [
                "0",
                "0.0",
                "0.00",
                "0.000",
                "0.0000",
                "-0",
                "-0.0",
                "-0.00",
                "-0.000",
                "0%",
                "0.0%",
                "0.00%",
                "0.000%",
                "-0%",
                "-0.0%",
                "-0.00%",
            ]

            if str_val in simple_zeros:
                return True

            # Try to extract number
            try:
                clean_str = re.sub(r"[^\d\.\-\+Ee]", "", str_val)
                if not clean_str:
                    return False
                float_val = float(clean_str)
                return abs(float_val) < 1e-10
            except (ValueError, TypeError):
                # Regex patterns for edge cases
                patterns = [
                    r"^0+$",
                    r"^0+\.0+$",
                    r"^-0+\.0+$",
                    r"^0+\.0*[%]?$",
                    r"^-0+\.0*[%]?$",
                    r"^0+\.0*e[+-]?\d+$",
                ]
                for pattern in patterns:
                    if re.match(pattern, str_val, re.IGNORECASE):
                        return True

                # Check if only zeros
                digits_only = re.sub(r"[^\d]", "", str_val)
                if digits_only and all(c == "0" for c in digits_only):
                    return True

        return False

    def detect_failed_kpis(self, file_path: Path) -> List[Tuple[int, str]]:
        """Detect failed KPIs from source file"""
        failed_kpis = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)

            if "Cluster & Tower" in wb.sheetnames:
                sheet = wb["Cluster & Tower"]
            elif "SSH Achievement" in wb.sheetnames:
                sheet = wb["SSH Achievement"]
            else:
                wb.close()
                return failed_kpis

            print("  🔍 Scanning for failed KPIs...")

            # Check KPI cells: M, P, S, V (rows 11-24)
            for row in range(11, 25):
                for col_letter, band_name in self.band_config.columns.items():
                    cell = sheet[f"{col_letter}{row}"]

                    if cell.value is None:
                        continue

                    if self.is_cell_zero_value(cell):
                        failed_kpis.append((row, band_name))
                        print(
                            f"    ⚠ Failed KPI: {col_letter}{row} = '{cell.value}' → {band_name}"
                        )

            wb.close()
            print(f"  📊 Total failed KPIs detected: {len(failed_kpis)}")

        except Exception as e:
            print(f"  ⚠ Error detecting failed KPIs: {e}")

        return failed_kpis

    @staticmethod
    def extract_cluster_tower(text: str) -> Tuple[str, str]:
        """Extract cluster and tower from text"""
        cluster_match = re.search(r"Cluster\s*:\s*([^T]+?)Tower", text)
        tower_match = re.search(r"Tower\s*:\s*([^\s(]+)", text)

        cluster = cluster_match.group(1).strip() if cluster_match else "Unknown"
        tower = tower_match.group(1).strip() if tower_match else "Unknown"

        return cluster, tower

    def create_sector_subplots(self, ax, sectors: List, num_sectors: int) -> List:
        """Create inner axes for sector-based charts"""
        axes = []
        gap = 0.06
        width = (
            (0.88 - (num_sectors - 1) * gap) / num_sectors if num_sectors > 1 else 0.88
        )

        for i in range(num_sectors):
            left = 0.06 + i * (width + gap)
            inner_ax = ax.inset_axes([left, 0.12, width, 0.76])
            inner_ax.set_frame_on(False)
            for spine in inner_ax.spines.values():
                spine.set_visible(False)
            axes.append(inner_ax)

        return axes

    def plot_metric_by_sector(
        self,
        ax,
        data: pd.DataFrame,
        metric_col: str,
        band_col: str,
        date_col: str,
        ylabel: str,
        ylim_min: Optional[float] = None,
        ylim_max: Optional[float] = None,
    ):
        """Generic plotting function for sector-based metrics"""
        sectors = data["newta_sector_name"].dropna().unique()
        if len(sectors) == 0:
            ax.text(
                0.5,
                0.5,
                "No sector data",
                ha="center",
                va="center",
                transform=ax.transAxes,
                fontsize=12,
                style="italic",
            )
            return

        num_sectors = min(3, len(sectors))
        inner_axes = self.create_sector_subplots(ax, sectors, num_sectors)

        for idx, sector in enumerate(sectors[:num_sectors]):
            inner_ax = inner_axes[idx]
            sector_data = data[data["newta_sector_name"] == sector]
            bands = sector_data[band_col].unique()

            for band_idx, band in enumerate(bands):
                band_data = sector_data[sector_data[band_col] == band].sort_values(
                    date_col
                )
                if band_data.empty:
                    continue

                band_str = str(int(band)) if pd.notna(band) else "Unknown"
                color = self.color_palette[band_idx % len(self.color_palette)]

                inner_ax.plot(
                    band_data[date_col],
                    band_data[metric_col],
                    marker="o",
                    label=f"L{band_str}",
                    color=color,
                    linewidth=2,
                    markersize=5,
                )

            inner_ax.set_title(
                f"Sector {sector}", fontsize=12, fontweight="bold", pad=10
            )
            inner_ax.set_xlabel("Date", fontsize=10)
            inner_ax.set_ylabel(ylabel, fontsize=10)
            inner_ax.legend(fontsize=8, loc="best", frameon=False)
            inner_ax.grid(True, linewidth=1.2, alpha=0.8, linestyle="-", color="gray")
            inner_ax.tick_params(axis="x", rotation=45, labelsize=9)
            inner_ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))

            if ylim_min is not None:
                inner_ax.set_ylim(bottom=ylim_min)
            if ylim_max is not None:
                inner_ax.set_ylim(top=ylim_max)

            if num_sectors > 1:
                inner_ax.set_facecolor("#f9f9f9")

    def plot_timing_advance(self, ax, ta_df: pd.DataFrame):
        """Plot Timing Advance with bars and CDF"""
        sectors = ta_df["newta_sector_name"].unique()
        if len(sectors) == 0:
            return

        num_sectors = min(3, len(sectors))
        inner_axes = self.create_sector_subplots(ax, sectors, num_sectors)

        distance_labels = [
            "0-78",
            "78-234",
            "234-390",
            "390-546",
            "546-702",
            "702-858",
            "858-1014",
            "1014-1560",
            "1560-2106",
            "2106-2652",
            "2652-3120",
            "3120-3900",
            "3900-6318",
            "6318-10062",
            "10062-13962",
            "13962-20000",
        ]
        x_pos = np.arange(len(distance_labels))

        for idx, sector in enumerate(sectors[:num_sectors]):
            inner_ax = inner_axes[idx]
            sector_data = ta_df[ta_df["newta_sector_name"] == sector]
            bands = sorted(
                [b for b in sector_data["newta_band"].unique() if pd.notna(b)]
            )
            bar_width = 0.8 / len(bands) if bands else 0.8

            for band_idx, band in enumerate(bands):
                band_data = sector_data[sector_data["newta_band"] == band].iloc[0]
                band_str = str(int(band))
                color = self.band_config.colors.get(band_str, "#95A5A6")

                # Distance distribution
                distance_cols = [
                    f"newta_{d}_m"
                    for d in [
                        "0_78",
                        "78_234",
                        "234_390",
                        "390_546",
                        "546_702",
                        "702_858",
                        "858_1014",
                        "1014_1560",
                        "1560_2106",
                        "2106_2652",
                        "2652_3120",
                        "3120_3900",
                        "3900_6318",
                        "6318_10062",
                        "10062_13962",
                        "13962_20000",
                    ]
                ]
                values = [
                    band_data[col]
                    if col in band_data and pd.notna(band_data[col])
                    else 0
                    for col in distance_cols
                ]

                pos = x_pos + band_idx * bar_width - (len(bands) - 1) * bar_width / 2
                inner_ax.bar(
                    pos,
                    values,
                    width=bar_width,
                    color=color,
                    alpha=0.7,
                    label=f"L{band_str}",
                )

            # Twin axis for CDF
            ax2 = inner_ax.twinx()
            for band_idx, band in enumerate(bands):
                band_data = sector_data[sector_data["newta_band"] == band].iloc[0]
                band_str = str(int(band))
                color = self.band_config.colors.get(band_str, "#95A5A6")

                cdf_cols = [
                    "newta_78",
                    "newta_234",
                    "newta_390",
                    "newta_546",
                    "newta_702",
                    "newta_858",
                    "newta_1014",
                    "newta_1560",
                    "newta_2106",
                    "newta_2652",
                    "newta_3120",
                    "newta_3900",
                    "newta_6318",
                    "newta_10062",
                    "newta_13962",
                    "newta_20000",
                ]
                cdf_values = [
                    band_data[col]
                    if col in band_data and pd.notna(band_data[col])
                    else 0
                    for col in cdf_cols
                ]

                ax2.plot(
                    x_pos,
                    cdf_values,
                    marker="o",
                    color=color,
                    linewidth=2,
                    label=f"L{band_str} CDF",
                )

            ax2.axhline(y=90, color="red", linestyle="--", linewidth=1.5, alpha=0.7)
            ax2.set_ylim(0, 105)
            ax2.grid(True, alpha=0.2, linestyle="--")

            inner_ax.set_title(
                f"Sector {sector}", fontsize=12, fontweight="bold", pad=10
            )
            inner_ax.set_xlabel("Distance (m)", fontsize=10)
            inner_ax.set_ylabel("Number of Samples", fontsize=10)
            inner_ax.set_xticks(x_pos)
            inner_ax.set_xticklabels(
                distance_labels, rotation=45, ha="right", fontsize=8
            )
            inner_ax.legend(loc="upper left", fontsize=8, ncol=2, frameon=False)
            ax2.legend(loc="upper right", fontsize=8, frameon=False)

            if num_sectors > 1:
                inner_ax.set_facecolor("#f9f9f9")

    def create_combined_chart(
        self,
        ta_df: pd.DataFrame,
        wd_merged: Optional[pd.DataFrame],
        bh_merged: Optional[pd.DataFrame],
        cluster: str,
        tower: str,
    ) -> Optional[bytes]:
        """Create combined chart with 4 rows"""
        if (
            ta_df.empty
            and (wd_merged is None or wd_merged.empty)
            and (bh_merged is None or bh_merged.empty)
        ):
            return None

        fig = plt.figure(figsize=(22, 28))
        gs = fig.add_gridspec(4, 1, hspace=0.4)

        # Row 1: CQI
        ax1 = fig.add_subplot(gs[0])
        ax1.set_frame_on(False)
        ax1.set_xticks([])
        ax1.set_yticks([])
        for spine in ax1.spines.values():
            spine.set_visible(False)

        if bh_merged is not None and not bh_merged.empty:
            self.plot_metric_by_sector(
                ax1,
                bh_merged,
                "cqi",
                "newbh_cell_fdd_band",
                "newbh_date",
                "Average CQI",
                0,
                15,
            )
            ax1.set_title("CQI - Busy Hour", fontsize=16, fontweight="bold", pad=20)
        else:
            ax1.text(
                0.5,
                0.5,
                "No CQI Data Available",
                ha="center",
                va="center",
                transform=ax1.transAxes,
                fontsize=12,
                style="italic",
            )
            ax1.set_title("CQI - Busy Hour", fontsize=16, fontweight="bold", pad=20)

        # Row 2: Spectral Efficiency
        ax2 = fig.add_subplot(gs[1])
        ax2.set_frame_on(False)
        ax2.set_xticks([])
        ax2.set_yticks([])
        for spine in ax2.spines.values():
            spine.set_visible(False)

        if wd_merged is not None and not wd_merged.empty:
            self.plot_metric_by_sector(
                ax2,
                wd_merged,
                "spectral_eff",
                "newwd_cell_fdd_band",
                "newwd_date",
                "Spectral Efficiency",
                0,
                None,
            )
            ax2.set_title(
                "Spectral Efficiency - Daily", fontsize=16, fontweight="bold", pad=20
            )
        else:
            ax2.text(
                0.5,
                0.5,
                "No Spectral Efficiency Data Available",
                ha="center",
                va="center",
                transform=ax2.transAxes,
                fontsize=12,
                style="italic",
            )
            ax2.set_title(
                "Spectral Efficiency - Daily", fontsize=16, fontweight="bold", pad=20
            )

        # Row 3: QPSK
        ax3 = fig.add_subplot(gs[2])
        ax3.set_frame_on(False)
        ax3.set_xticks([])
        ax3.set_yticks([])
        for spine in ax3.spines.values():
            spine.set_visible(False)

        if bh_merged is not None and not bh_merged.empty:
            self.plot_metric_by_sector(
                ax3,
                bh_merged,
                "qpsk",
                "newbh_cell_fdd_band",
                "newbh_date",
                "QPSK Rate (%)",
                0,
                100,
            )
            ax3.set_title(
                "QPSK Rate - Busy Hour", fontsize=16, fontweight="bold", pad=20
            )
        else:
            ax3.text(
                0.5,
                0.5,
                "No QPSK Data Available",
                ha="center",
                va="center",
                transform=ax3.transAxes,
                fontsize=12,
                style="italic",
            )
            ax3.set_title(
                "QPSK Rate - Busy Hour", fontsize=16, fontweight="bold", pad=20
            )

        # Row 4: Timing Advance
        ax4 = fig.add_subplot(gs[3])
        ax4.set_frame_on(False)
        ax4.set_xticks([])
        ax4.set_yticks([])
        for spine in ax4.spines.values():
            spine.set_visible(False)

        if not ta_df.empty:
            self.plot_timing_advance(ax4, ta_df)
            ax4.set_title("Timing Advance", fontsize=14, fontweight="bold", pad=20)
        else:
            ax4.text(
                0.5,
                0.5,
                "No Timing Advance Data Available",
                ha="center",
                va="center",
                transform=ax4.transAxes,
                fontsize=12,
                style="italic",
            )
            ax4.set_title("Timing Advance", fontsize=14, fontweight="bold", pad=20)

        plt.tight_layout(rect=[0, 0, 1, 1])
        buf = BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()

    def create_failed_kpi_charts(
        self,
        ta_df: pd.DataFrame,
        wd_merged: Optional[pd.DataFrame],
        bh_merged: Optional[pd.DataFrame],
        kqi_df: Optional[pd.DataFrame],
        failed_kpis: List[Tuple[int, str]],
        cluster: str,
        tower: str,
    ) -> Optional[bytes]:
        """Generate charts for failed KPIs"""
        if not failed_kpis:
            print("  ℹ No failed KPIs detected → skipping failed KPI charts")
            return None

        print(f"  📊 Generating {len(failed_kpis)} Failed KPI chart(s)...")

        height_per_chart = 5.5
        fig = plt.figure(figsize=(22, max(6, height_per_chart * len(failed_kpis))))
        gs = fig.add_gridspec(len(failed_kpis), 1, hspace=0.6)

        chart_idx = 0
        plotted_any = False

        data_sources = {
            "bh": bh_merged,
            "wd": wd_merged,
            "kqi": kqi_df,
        }

        for row_num, band_name in failed_kpis:
            ax = fig.add_subplot(gs[chart_idx])
            ax.set_frame_on(False)
            ax.set_xticks([])
            ax.set_yticks([])
            for spine in ax.spines.values():
                spine.set_visible(False)

            config = self.kpi_configs.get(row_num)
            if not config:
                ax.text(
                    0.5,
                    0.5,
                    f"KPI Row {row_num} not supported",
                    ha="center",
                    va="center",
                    transform=ax.transAxes,
                    fontsize=12,
                    color="gray",
                )
                chart_idx += 1
                continue

            source_df = data_sources.get(config.source_df_name)

            if row_num in [18, 19]:
                kpi_name = self.kpi_mapping.get(row_num, f"Row {row_num}")
                ax.set_title(
                    f"{kpi_name}",
                    fontsize=16,
                    fontweight="bold",
                    pad=20,
                )
                
                if source_df is None or source_df.empty:
                    ax.text(
                        0.5,
                        0.5,
                        "No data available",
                        ha="center",
                        va="center",
                        transform=ax.transAxes,
                        fontsize=12,
                        color="gray",
                    )
                else:
                    inner_ax = ax.inset_axes([0.06, 0.12, 0.88, 0.76])
                    plot_data = source_df[[config.date_column, config.column_name]].dropna()
                    
                    if plot_data.empty:
                        ax.text(
                            0.5,
                            0.5,
                            "No valid data points",
                            ha="center",
                            va="center",
                            transform=ax.transAxes,
                            fontsize=12,
                            color="gray",
                        )
                    else:
                        # Sort by date
                        plot_data = plot_data.sort_values(config.date_column)
                        
                        # Plot the line
                        inner_ax.plot(
                            plot_data[config.date_column],
                            plot_data[config.column_name],
                            marker="o",
                            linewidth=3,
                            markersize=8,
                            color="#f00dd2",
                            label="Overall Latency",
                        )
                        
                        if row_num == 19:
                            inner_ax.axhline(
                                y=120,
                                color="#FFA500",
                                linestyle="-.",
                                linewidth=2.5,
                                alpha=0.8,
                                label="Threshold 120ms"
                            )
                            inner_ax.axhline(
                                y=200,
                                color="#FF4500",
                                linestyle="-.",
                                linewidth=2.5,
                                alpha=0.8,
                                label="Critical 200ms"
                            )
                        elif row_num == 18:
                            inner_ax.axhline(
                                y=1.0,
                                color="#FFA500",
                                linestyle="-.",
                                linewidth=2.5,
                                alpha=0.8,
                                label="Threshold 1%"
                            )
                            inner_ax.axhline(
                                y=2.0,
                                color="#FF4500",
                                linestyle="-.",
                                linewidth=2.5,
                                alpha=0.8,
                                label="Critical 2%"
                            )
                        
                        # Formatting
                        inner_ax.set_ylabel(f"{config.title} ({config.unit})", fontsize=12, fontweight="bold")
                        inner_ax.set_xlabel("Date", fontsize=12, fontweight="bold")
                        inner_ax.grid(True, linewidth=1.2, alpha=0.8, linestyle="-", color="gray")
                        inner_ax.tick_params(axis="x", rotation=45, labelsize=10)
                        inner_ax.tick_params(axis="y", labelsize=10)
                        inner_ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
                        inner_ax.xaxis.set_major_locator(mdates.AutoDateLocator())
                        inner_ax.legend(fontsize=11, loc="best", frameon=True, shadow=True)
                        
                        # Set y-axis limits
                        if config.ylim_min is not None:
                            inner_ax.set_ylim(bottom=config.ylim_min)
                        if config.ylim_max is not None:
                            inner_ax.set_ylim(top=config.ylim_max)
                        else:
                            # Auto-scale with some padding
                            y_min = plot_data[config.column_name].min()
                            y_max = plot_data[config.column_name].max()
                            
                            # Adjust y_max to include threshold lines
                            if row_num == 19:  # Latency
                                y_max = max(y_max, 200) + 50
                            elif row_num == 18:  # Packet Loss
                                y_max = max(y_max, 2.0) + 0.5
                            
                            y_range = y_max - y_min
                            inner_ax.set_ylim(
                                bottom=max(0, y_min - y_range * 0.05),
                                top=y_max + y_range * 0.05
                            )
                        
                        # Set spine visibility
                        for spine in inner_ax.spines.values():
                            spine.set_visible(True)
                            spine.set_linewidth(1.5)
                        
                        plotted_any = True

            # Sector-based KPIs
            else:
                if source_df is None or source_df.empty:
                    ax.text(
                        0.5,
                        0.5,
                        "No data available",
                        ha="center",
                        va="center",
                        transform=ax.transAxes,
                        fontsize=12,
                        color="gray",
                    )
                else:
                    allowed_bands = self.band_config.values.get(band_name, [])
                    filtered_df = source_df.copy()

                    if filtered_df.empty:
                        ax.text(
                            0.5,
                            0.5,
                            f"No data for {band_name}",
                            ha="center",
                            va="center",
                            transform=ax.transAxes,
                            fontsize=12,
                            color="gray",
                        )
                    else:
                        sectors = filtered_df["newta_sector_name"].dropna().unique()
                        if len(sectors) == 0:
                            ax.text(
                                0.5,
                                0.5,
                                "No sector data",
                                ha="center",
                                va="center",
                                transform=ax.transAxes,
                                fontsize=12,
                                color="gray",
                            )
                        else:
                            num_sectors = min(3, len(sectors))
                            inner_axes = self.create_sector_subplots(
                                ax, sectors, num_sectors
                            )

                            for idx, sector in enumerate(sectors[:num_sectors]):
                                inner_ax = inner_axes[idx]
                                sector_data = filtered_df[
                                    filtered_df["newta_sector_name"] == sector
                                ]
                                bands = sector_data[config.band_column].unique()

                                for band_idx, band in enumerate(bands):
                                    band_data = sector_data[
                                        sector_data[config.band_column] == band
                                    ].sort_values(config.date_column)
                                    if band_data.empty:
                                        continue

                                    band_str = (
                                        str(int(band)) if pd.notna(band) else "Unknown"
                                    )

                                    # Highlight failed band
                                    if int(band) in allowed_bands:
                                        color = "#d32f2f"
                                        linewidth = 4
                                        alpha = 1.0
                                    else:
                                        color = self.color_palette[
                                            band_idx % len(self.color_palette)
                                        ]
                                        linewidth = 2
                                        alpha = 0.6

                                    inner_ax.plot(
                                        band_data[config.date_column],
                                        band_data[config.column_name],
                                        marker="o",
                                        label=f"L{band_str}",
                                        color=color,
                                        linewidth=linewidth,
                                        markersize=5,
                                        alpha=alpha,
                                    )

                                inner_ax.set_title(
                                    f"Sector {sector}",
                                    fontsize=12,
                                    fontweight="bold",
                                    pad=10,
                                )
                                inner_ax.set_xlabel("Date", fontsize=10)
                                inner_ax.set_ylabel(
                                    f"{config.title} ({config.unit})", fontsize=10
                                )
                                inner_ax.legend(fontsize=8, loc="best", frameon=False)
                                inner_ax.grid(
                                    True,
                                    linewidth=1.2,
                                    alpha=0.8,
                                    linestyle="-",
                                    color="gray",
                                )
                                inner_ax.tick_params(axis="x", rotation=45, labelsize=9)
                                inner_ax.xaxis.set_major_formatter(
                                    mdates.DateFormatter("%m/%d")
                                )

                                if config.ylim_min is not None:
                                    inner_ax.set_ylim(bottom=config.ylim_min)
                                if config.ylim_max is not None:
                                    inner_ax.set_ylim(top=config.ylim_max)

                                if num_sectors > 1:
                                    inner_ax.set_facecolor("#f9f9f9")

                            plotted_any = True

                kpi_name = self.kpi_mapping.get(row_num, f"Row {row_num}")
                ax.set_title(
                    f"{kpi_name} - {band_name}", fontsize=16, fontweight="bold", pad=20
                )

            chart_idx += 1

        if not plotted_any:
            print("  ⚠ No chart was successfully plotted → skipping image")
            plt.close(fig)
            return None

        plt.tight_layout(rect=[0, 0, 1, 1])
        buf = BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        buf.seek(0)
        print("  ✓ Failed KPI charts generated successfully")
        return buf.getvalue()

    def copy_range_with_excel(
        self, source_file: Path, template_file: Path, output_file: Path
    ) -> bool:
        """Copy range as picture using PowerShell/Excel COM (Windows only)"""
        ps_script = f'''
$ErrorActionPreference = "Stop"
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false

    $sourceWb = $excel.Workbooks.Open("{source_file}")
    $sourceWs = $sourceWb.Worksheets.Item("SSH Achievement")
    $range = $sourceWs.Range("B6:V34")
    $range.CopyPicture(1, 2)

    $templateWb = $excel.Workbooks.Open("{template_file}")

    if ($templateWb.Worksheets | Where-Object {{ $_.Name -eq "SSH Achievement" }}) {{
        $sshWs = $templateWb.Worksheets.Item("SSH Achievement")
        $sshWs.Activate()
        $pasteRangeSSH = $sshWs.Range("C9")
        $pasteRangeSSH.Select()
        $sshWs.Paste()
    }}

    if ($templateWb.Worksheets | Where-Object {{ $_.Name -eq "Justification" }}) {{
        $justWs = $templateWb.Worksheets.Item("Justification")
        $justWs.Activate()
        $pasteRangeJust = $justWs.Range("C9")
        $pasteRangeJust.Select()
        $justWs.Paste()
    }}

    $templateWb.SaveAs("{output_file}", 51)
    $templateWb.Close($false)
    $sourceWb.Close($false)
    $excel.Quit()

    [System.Runtime.Interopservices.Marshal]::ReleaseComObject($sourceWs) | Out-Null
    [System.Runtime.Interopservices.Marshal]::ReleaseComObject($sourceWb) | Out-Null
    [System.Runtime.Interopservices.Marshal]::ReleaseComObject($templateWb) | Out-Null
    [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
    [System.GC]::Collect()
    [System.GC]::WaitForPendingFinalizers()

    exit 0
}} catch {{
    Write-Host "ERROR: $_"
    exit 1
}}
'''
        ps_file = self.output_folder / "temp_copy.ps1"
        try:
            with open(ps_file, "w", encoding="utf-8") as f:
                f.write(ps_script)

            print("  📝 Running PowerShell script...")
            result = subprocess.run(
                ["powershell", "-ExecutionPolicy", "Bypass", "-File", str(ps_file)],
                capture_output=True,
                text=True,
                encoding="utf-8",
                timeout=120,
            )

            if result.returncode == 0 and output_file.exists():
                print(f"  ✓ Excel automation successful")
                return True
            else:
                print(f"  ✗ PowerShell failed (code {result.returncode})")
                return False

        except subprocess.TimeoutExpired:
            print("  ✗ PowerShell script timed out")
            return False
        except Exception as e:
            print(f"  ✗ PowerShell execution error: {e}")
            return False
        finally:
            ps_file.unlink(missing_ok=True)

    def apply_conditional_values(self, ws):
        """Apply conditional values based on cell checks"""
        k_values = [
            "99.00",
            "98.00",
            "1.00",
            "97.00",
            "8.00",
            "50.00",
            "1.10",
            "1.00",
            "120.00",
            "5.00",
            "1.00",
            "20.00",
            "-105.00",
            "30.00",
        ]
        n_values = [
            "99.00",
            "98.00",
            "1.00",
            "97.00",
            "9.00",
            "40.00",
            "1.50",
            "1.00",
            "120.00",
            "10.00",
            "1.50",
            "35.00",
            "-105.00",
            "30.00",
            "98.00",
            "98.00",
            "5.00",
        ]
        q_values = [
            "99.00",
            "98.00",
            "1.00",
            "97.00",
            "9.00",
            "40.00",
            "1.70",
            "1.00",
            "120.00",
            "10.00",
            "1.50",
            "35.00",
            "-105.00",
            "30.00",
        ]
        t_values = [
            "99.00",
            "98.00",
            "1.00",
            "97.00",
            "10.00",
            "40.00",
            "1.90",
            "1.00",
            "120.00",
            "10.00",
            "1.50",
            "35.00",
            "-105.00",
            "30.00",
        ]

        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

        conditions = [
            ("M28", k_values, "K", range(11, 25)),
            ("P28", n_values, "N", range(11, 28)),
            ("S28", q_values, "Q", range(11, 25)),
            ("V28", t_values, "T", range(11, 25)),
        ]

        for check_cell, values, col, rows in conditions:
            try:
                if ws[check_cell].value and float(ws[check_cell].value) > 0:
                    for idx, i in enumerate(rows):
                        if idx < len(values):
                            cell = ws[f"{col}{i}"]
                            cell.value = values[idx]
                            cell.fill = white_fill
                            cell.number_format = "@"
            except Exception as e:
                print(f"  ⚠ Warning {check_cell}: {e}")

    def write_failed_kpi_list(self, wb, failed_kpis: List[Tuple[int, str]]):
        """Write failed KPI list to Justification sheet"""
        if "Justification" not in wb.sheetnames or not failed_kpis:
            return

        just_sheet = wb["Justification"]
        start_row = 247

        just_sheet[f"F{start_row - 1}"].value = "KPI Failed List:"
        just_sheet[f"F{start_row - 1}"].font = Font(bold=True, size=14, color="2A2727")

        for idx, (row_num, band_name) in enumerate(failed_kpis):
            kpi_name = self.kpi_mapping.get(row_num, f"Unknown KPI (Row {row_num})")
            display_name = (
                f"{kpi_name} {band_name}" if band_name != "Overall" else kpi_name
            )
            cell = just_sheet[f"F{start_row + idx}"]
            cell.value = display_name
            cell.font = Font(color="2A2727", bold=True, size=11)

    def add_wd_latest_data(self, report_path: Path, wd_df: pd.DataFrame):
        """Add latest WD data to SSH Achievement sheet"""
        try:
            if wd_df is None or wd_df.empty:
                print("  ℹ No WD data to add")
                return

            cols_needed = [
                "newwd_date",
                "newwd_operator",
                "newwd_moentity",
                "newwd_enodeb_fdd_msc",
                "newwd_cell_fdd_system",
                "newwd_cell_fdd_txrxmode",
                "newwd_cell_fdd_vendor",
                "newwd_cell_fdd_band",
            ]

            available_cols = [col for col in cols_needed if col in wd_df.columns]
            wd_filtered = wd_df[available_cols].copy()

            if "newwd_date" in wd_filtered.columns:
                wd_filtered["date_str"] = wd_filtered["newwd_date"].dt.strftime(
                    "%Y-%m-%d"
                )
                max_date = wd_filtered["date_str"].max()
                latest_data = wd_filtered[wd_filtered["date_str"] == max_date]
            else:
                latest_data = wd_filtered

            latest_data = latest_data.drop_duplicates().drop(
                columns=["date_str"], errors="ignore"
            )
            latest_data = latest_data[cols_needed]

            if latest_data.empty:
                print("  ℹ No unique latest WD data found")
                return

            wb = openpyxl.load_workbook(report_path)

            if "SSH Achievement" not in wb.sheetnames:
                print("  ⚠ Sheet 'SSH Achievement' not found for WD data")
                wb.close()
                return

            sheet = wb["SSH Achievement"]
            start_row, start_col = 53, 3

            for r_idx, row in enumerate(
                latest_data.itertuples(index=False, name=None), start=start_row
            ):
                for c_idx, value in enumerate(row, start=start_col):
                    cell = sheet.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    if c_idx == start_col and value is not None:
                        try:
                            cell.number_format = "YYYY-MM-DD"
                        except:
                            pass

            # Apply borders
            num_rows = len(latest_data)
            end_row = start_row + num_rows - 1
            end_col_letter = openpyxl.utils.get_column_letter(
                start_col + len(cols_needed) - 1
            )

            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for row in sheet[f"C{start_row}:{end_col_letter}{end_row}"]:
                for cell in row:
                    cell.border = border

            wb.save(report_path)
            wb.close()
            print(f"  ✓ Added {num_rows} WD records to SSH Achievement")

        except Exception as e:
            print(f"  ⚠ Failed to add WD data: {e}")

    def process_file(self, file_path: Path) -> Optional[Path]:
        """Process a single file and generate report"""
        print(f"\nProcessing: {file_path.name}")

        try:
            # Step 1: Extract metadata and detect failed KPIs
            wb = openpyxl.load_workbook(file_path)

            if "Cluster & Tower" in wb.sheetnames:
                source_sheet = wb["Cluster & Tower"]
            elif "SSH Achievement" in wb.sheetnames:
                source_sheet = wb["SSH Achievement"]
            else:
                print("  ✗ No valid sheet found")
                wb.close()
                return None

            b6_value = source_sheet["B6"].value or ""
            cluster, tower = self.extract_cluster_tower(str(b6_value))
            print(f"  ✓ Cluster: {cluster} | Tower: {tower}")
            wb.close()

            # Detect failed KPIs
            failed_kpis = self.detect_failed_kpis(file_path)

            # Step 2: Prepare source file
            wb = openpyxl.load_workbook(file_path)
            if "Cluster & Tower" in wb.sheetnames:
                ws = wb["Cluster & Tower"]
                ws.title = "SSH Achievement"

            sheets_to_remove = [s for s in wb.sheetnames if s != "SSH Achievement"]
            for sheet in sheets_to_remove:
                wb.remove(wb[sheet])

            self.apply_conditional_values(wb["SSH Achievement"])

            timestamp = datetime.now().strftime("%H%M%S")
            modified_source = (
                self.output_folder / f"temp_{file_path.stem}_{timestamp}.xlsx"
            )
            wb.save(modified_source)
            wb.close()

            time.sleep(1)

            # Step 3: Query and prepare data
            ta_df, wd_df, bh_df, kqi_df = self.query_data(tower)
            ta_df, wd_merged, bh_merged = self.merge_data(ta_df, wd_df, bh_df)
            self.compute_derived_metrics(wd_merged, bh_merged, kqi_df)

            # Step 4: Generate report
            timestamp_full = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = (
                f"SSH_Achievement_Report_NS_{cluster}_{tower}_{timestamp_full}.xlsx"
            )
            output_path = self.output_folder / output_filename

            # Try Excel automation first
            print("  ⚡ Attempting Excel automation...")
            success = self.copy_range_with_excel(
                modified_source, self.template_path, output_path
            )

            if not success or not output_path.exists():
                print("  ⚡ Falling back to Python-only method...")
                result = self.create_report_python_only(
                    file_path,
                    cluster,
                    tower,
                    ta_df,
                    wd_merged,
                    bh_merged,
                    kqi_df,
                    wd_df,
                    failed_kpis,
                    output_path,
                )
            else:
                self.finalize_report(
                    output_path,
                    cluster,
                    tower,
                    ta_df,
                    wd_merged,
                    bh_merged,
                    kqi_df,
                    wd_df,
                    failed_kpis,
                )
                result = output_path

            # Cleanup
            modified_source.unlink(missing_ok=True)

            if result and result.exists():
                print(f"  ✓ Successfully created: {result.name}")
                return result
            else:
                print("  ✗ Final report not created")
                return None

        except Exception as e:
            print(f"  ✗ Error processing {file_path.name}: {e}")
            import traceback

            traceback.print_exc()
            return None

    def create_report_python_only(
        self,
        original_file: Path,
        cluster: str,
        tower: str,
        ta_df: pd.DataFrame,
        wd_merged: Optional[pd.DataFrame],
        bh_merged: Optional[pd.DataFrame],
        kqi_df: Optional[pd.DataFrame],
        wd_df: pd.DataFrame,
        failed_kpis: List[Tuple[int, str]],
        suggested_path: Path,
    ) -> Optional[Path]:
        """Create report using Python only (no Excel automation)"""
        print("  ⚡ Using Python-only method...")

        try:
            template_wb = openpyxl.load_workbook(self.template_path)
            template_ws = (
                template_wb["SSH Achievement"]
                if "SSH Achievement" in template_wb.sheetnames
                else template_wb.active
            )

            source_wb = openpyxl.load_workbook(original_file)
            source_ws = source_wb["SSH Achievement"]

            # Copy range B6:V34
            for row_idx, row in enumerate(source_ws["B6:V34"], start=8):
                for cell in row:
                    target_cell = template_ws.cell(row=row_idx, column=cell.column + 1)
                    target_cell.value = cell.value
                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.border = copy(cell.border)
                        target_cell.fill = copy(cell.fill)
                        target_cell.number_format = cell.number_format
                        target_cell.alignment = copy(cell.alignment)

            source_wb.close()

            # Update cluster & tower
            template_ws["E34"] = cluster
            template_ws["E34"].font = Font(bold=True, size=12)
            template_ws["E35"] = tower
            template_ws["E35"].font = Font(bold=True, size=12)

            # Add charts
            if "Justification" in template_wb.sheetnames:
                just_sheet = template_wb["Justification"]

                chart_bytes = self.create_combined_chart(
                    ta_df, wd_merged, bh_merged, cluster, tower
                )
                if chart_bytes:
                    img = XLImage(BytesIO(chart_bytes))
                    img.anchor = "B46"
                    just_sheet.add_image(img)
                    print("  ✓ Added Combined Chart")

                if failed_kpis:
                    failed_chart = self.create_failed_kpi_charts(
                        ta_df, wd_merged, bh_merged, kqi_df, failed_kpis, cluster, tower
                    )
                    if failed_chart:
                        img_failed = XLImage(BytesIO(failed_chart))
                        img_failed.anchor = "C250"
                        just_sheet.add_image(img_failed)
                        print(f"  ✓ Added {len(failed_kpis)} Failed KPI charts")

                self.write_failed_kpi_list(template_wb, failed_kpis)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt_name = (
                f"SSH_Achievement_Report_NS_{cluster}_{tower}_{timestamp}_ALT.xlsx"
            )
            alt_path = self.output_folder / alt_name
            template_wb.save(alt_path)
            template_wb.close()

            self.add_wd_latest_data(alt_path, wd_df)
            print(f"  ✓ Created report: {alt_name}")
            return alt_path

        except Exception as e:
            print(f"  ✗ Python-only method failed: {e}")
            import traceback

            traceback.print_exc()
            return None

    def finalize_report(
        self,
        report_path: Path,
        cluster: str,
        tower: str,
        ta_df: pd.DataFrame,
        wd_merged: Optional[pd.DataFrame],
        bh_merged: Optional[pd.DataFrame],
        kqi_df: Optional[pd.DataFrame],
        wd_df: pd.DataFrame,
        failed_kpis: List[Tuple[int, str]],
    ):
        """Finalize report by adding charts and data"""
        try:
            wb = openpyxl.load_workbook(report_path)

            if "Justification" in wb.sheetnames:
                just_sheet = wb["Justification"]

                # Main chart
                main_chart = self.create_combined_chart(
                    ta_df, wd_merged, bh_merged, cluster, tower
                )
                if main_chart:
                    img = XLImage(BytesIO(main_chart))
                    img.anchor = "C49"
                    just_sheet.add_image(img)
                    print("  ✓ Added main chart")

                # Failed KPI charts
                if failed_kpis:
                    failed_chart = self.create_failed_kpi_charts(
                        ta_df, wd_merged, bh_merged, kqi_df, failed_kpis, cluster, tower
                    )
                    if failed_chart:
                        img_failed = XLImage(BytesIO(failed_chart))
                        img_failed.anchor = "C250"
                        just_sheet.add_image(img_failed)
                        print(f"  ✓ Added {len(failed_kpis)} failed KPI charts")

                self.write_failed_kpi_list(wb, failed_kpis)

            # Update metadata
            if "SSH Achievement" in wb.sheetnames:
                ssh_sheet = wb["SSH Achievement"]
                ssh_sheet["E34"] = cluster
                ssh_sheet["E34"].font = Font(bold=True, size=12)
                ssh_sheet["E35"] = tower
                ssh_sheet["E35"].font = Font(bold=True, size=12)

            wb.save(report_path)
            wb.close()

            self.add_wd_latest_data(report_path, wd_df)

        except Exception as e:
            print(f"  ⚠ Finalization error: {e}")
            import traceback

            traceback.print_exc()

    def list_xlsx_files(self) -> List[Path]:
        """List all Excel files in input folder"""
        xlsx_files = sorted(self.input_folder.glob("*.xlsx"))
        print(f"Found {len(xlsx_files)} Excel files:")
        for i, file in enumerate(xlsx_files, 1):
            print(f"  {i}. {file.name}")
        return xlsx_files

    def process_all_files(self):
        """Process all Excel files in input folder"""
        files = self.list_xlsx_files()
        if not files:
            print("No Excel files found in the input folder.")
            return

        print(f"\n{'=' * 60}")
        print("Starting batch processing...")
        print(f"{'=' * 60}")

        successful = 0
        failed = 0

        for file in files:
            result = self.process_file(file)
            if result:
                successful += 1
            else:
                failed += 1

        print(f"\n{'=' * 60}")
        print("Processing complete!")
        print(f"  ✓ Successful: {successful}")
        print(f"  ✗ Failed: {failed}")
        print(f"  Output folder: {self.output_folder}")
        print(f"{'=' * 60}")


def main():
    INPUT_FOLDER = "D:\\NEW SITE\\WORK\\REQ\\20260112\\"
    TEMPLATE_PATH = "./template.xlsx"
    OUTPUT_FOLDER = "D:\\NEW SITE\\WORK\\REQ\\20260112\\output_reports\\"
    DB_PATH = "./newdatabase.db"

    generator = SSHReportGenerator(INPUT_FOLDER, TEMPLATE_PATH, OUTPUT_FOLDER, DB_PATH)
    generator.process_all_files()


if __name__ == "__main__":
    main()
