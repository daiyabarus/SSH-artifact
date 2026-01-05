import os
import re
import sqlite3
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from copy import copy
from io import BytesIO
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
import subprocess
import time
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class SSHReportGenerator:
    def __init__(self, input_folder, template_path, output_folder, db_path):
        # Resolve all paths to absolute to avoid relative path issues in PowerShell/COM
        self.input_folder = Path(input_folder).resolve()
        self.template_path = Path(template_path).resolve()
        self.output_folder = Path(output_folder).resolve()
        self.db_path = Path(db_path).resolve()
        self.output_folder.mkdir(exist_ok=True, parents=True)
        
        # Color palettes
        self.color_palette = [
            "#080cec", "#ef17e8", "#52eb0c", "#f59e0b", "#10b981",
            "#06b6d4", "#ef4444", "#a855f7", "#14b8a6", "#f97316"
        ]
        
        self.band_colors = {
            "850": "#52eb0c",
            "1800": "#080cec",
            "2100": "#ef17e8",
            "2300F1": "#F39C12",
            "2300F2": "#9B59B6",
        }
    
    def connect_db(self):
        if not self.db_path.exists():
            raise FileNotFoundError(f"Database not found: {self.db_path}")
        return sqlite3.connect(self.db_path)
    
    def parse_date_flexible(self, date_str):
        if pd.isna(date_str):
            return None
        
        date_formats = [
            "%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", 
            "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(str(date_str).strip(), fmt)
            except ValueError:
                continue
        print(f"  ⚠ Could not parse date: {date_str}")
        return None

    def query_ta_data(self, conn, tower_id):
        query = """
        SELECT * FROM tbl_newta 
        WHERE newta_managed_element = ?
        ORDER BY newta_date DESC
        """
        df = pd.read_sql_query(query, conn, params=(tower_id,))
        return df
    
    def query_wd_data(self, conn, tower_id):
        query = """
        SELECT newwd_date, 
                newwd_operator, 
                newwd_moentity, 
                newwd_enodeb_fdd_msc, 
                newwd_cell_fdd_system, 
                newwd_cell_fdd_txrxmode, 
                newwd_cell_fdd_vendor, 
                newwd_cell_fdd_band,
                newwd_spectral_efficiency_dl_num, newwd_spectral_efficiency_dl_den
        FROM tbl_newwd
        WHERE newwd_enodeb_fdd_msc = ?
        ORDER BY newwd_date
        """
        df = pd.read_sql_query(query, conn, params=(tower_id,))
        df['newwd_date'] = df['newwd_date'].apply(self.parse_date_flexible)
        return df
    
    def query_bh_data(self, conn, tower_id):
        query = """
        SELECT newbh_date, newbh_moentity, 
               newbh_cell_average_cqi_num, newbh_cell_average_cqi_den,
               newbh_cell_qpsk_rate_num, newbh_cell_qpsk_rate_den,
               newbh_cell_fdd_band, newbh_enodeb_fdd_msc
        FROM tbl_newbh
        WHERE newbh_enodeb_fdd_msc = ?
        ORDER BY newbh_date
        """
        df = pd.read_sql_query(query, conn, params=(tower_id,))
        df['newbh_date'] = df['newbh_date'].apply(self.parse_date_flexible)
        return df
    
    def merge_data(self, ta_df, wd_df, bh_df):
        if ta_df.empty:
            return None, None, None
        
        wd_merged = None
        if not wd_df.empty:
            wd_merged = pd.merge(
                wd_df, 
                ta_df[['newta_eutrancell', 'newta_sector', 'newta_sector_name']],
                left_on='newwd_moentity',
                right_on='newta_eutrancell',
                how='left'
            )
        
        bh_merged = None
        if not bh_df.empty:
            bh_merged = pd.merge(
                bh_df,
                ta_df[['newta_eutrancell', 'newta_sector', 'newta_sector_name']],
                left_on='newbh_moentity',
                right_on='newta_eutrancell',
                how='left'
            )
        
        return ta_df, wd_merged, bh_merged

    def _clean_outer_axis(self, ax):
        """Remove all frames, ticks, and spines from outer axis"""
        ax.set_frame_on(False)
        ax.set_xticks([])
        ax.set_yticks([])
        for spine in ax.spines.values():
            spine.set_visible(False)

    def create_combined_chart(self, ta_df, wd_merged, bh_merged, cluster, tower):
        """Create a single figure with 4 charts arranged in 4 rows"""
        if ta_df.empty and (wd_merged is None or wd_merged.empty) and (bh_merged is None or bh_merged.empty):
            return None
        
        fig = plt.figure(figsize=(22, 28))
        gs = fig.add_gridspec(4, 1, hspace=0.4, wspace=0.1)
        
        # Row 1: CQI Chart
        ax1 = fig.add_subplot(gs[0])
        self._clean_outer_axis(ax1)
        if bh_merged is not None and not bh_merged.empty:
            self._plot_cqi_on_axis(ax1, bh_merged)
            ax1.set_title("CQI - Busy Hour", fontsize=14, fontweight='bold', pad=20)
        else:
            ax1.text(0.5, 0.5, "No CQI Data Available", ha='center', va='center', transform=ax1.transAxes, fontsize=12, style='italic')
            ax1.set_title("CQI - Busy Hour", fontsize=14, fontweight='bold', pad=20)
        
        # Row 2: Spectral Efficiency Chart
        ax2 = fig.add_subplot(gs[1])
        self._clean_outer_axis(ax2)
        if wd_merged is not None and not wd_merged.empty:
            self._plot_spectral_efficiency_on_axis(ax2, wd_merged)
            ax2.set_title("Spectral Efficiency - Daily", fontsize=14, fontweight='bold', pad=20)
        else:
            ax2.text(0.5, 0.5, "No Spectral Efficiency Data Available", ha='center', va='center', transform=ax2.transAxes, fontsize=12, style='italic')
            ax2.set_title("Spectral Efficiency - Daily", fontsize=14, fontweight='bold', pad=20)
        
        # Row 3: QPSK Chart
        ax3 = fig.add_subplot(gs[2])
        self._clean_outer_axis(ax3)
        if bh_merged is not None and not bh_merged.empty:
            self._plot_qpsk_on_axis(ax3, bh_merged)
            ax3.set_title("QPSK Rate - Busy Hour", fontsize=14, fontweight='bold', pad=20)
        else:
            ax3.text(0.5, 0.5, "No QPSK Data Available", ha='center', va='center', transform=ax3.transAxes, fontsize=12, style='italic')
            ax3.set_title("QPSK Rate - Busy Hour", fontsize=14, fontweight='bold', pad=20)
        
        # Row 4: Timing Advance Chart
        ax4 = fig.add_subplot(gs[3])
        self._clean_outer_axis(ax4)
        if not ta_df.empty:
            self._plot_timing_advance_on_axis(ax4, ta_df)
            ax4.set_title("Timing Advance", fontsize=14, fontweight='bold', pad=20)
        else:
            ax4.text(0.5, 0.5, "No Timing Advance Data Available", ha='center', va='center', transform=ax4.transAxes, fontsize=12, style='italic')
            ax4.set_title("Timing Advance", fontsize=14, fontweight='bold', pad=20)
        
        # main_title = f"Trend Charts \nCluster: {cluster} | Tower: {tower}"
        # fig.suptitle(main_title, fontsize=16, fontweight='bold', y=0.98)
        
        plt.tight_layout(rect=[0, 0, 1, 0.96])
        
        buf = BytesIO()
        fig.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()

    def _plot_cqi_on_axis(self, ax, bh_merged):
        bh_merged['cqi'] = np.where(
            bh_merged['newbh_cell_average_cqi_den'] > 0,
            bh_merged['newbh_cell_average_cqi_num'] / bh_merged['newbh_cell_average_cqi_den'],
            np.nan
        )
        
        sectors = bh_merged['newta_sector_name'].dropna().unique()
        if len(sectors) == 0:
            return
        
        num_sectors = min(3, len(sectors))
        axes = []
        
        # Create inner axes directly without intermediate container
        gap = 0.06
        width = (0.88 - (num_sectors - 1) * gap) / num_sectors if num_sectors > 1 else 0.88
        
        for i in range(num_sectors):
            left = 0.06 + i * (width + gap)
            inner_ax = ax.inset_axes([left, 0.12, width, 0.76])
            inner_ax.set_frame_on(False)
            for spine in inner_ax.spines.values():
                spine.set_visible(False)
            axes.append(inner_ax)
        
        if num_sectors == 1:
            axes = [ax.inset_axes([0.06, 0.12, 0.88, 0.76])]
            axes[0].set_frame_on(False)
            for spine in axes[0].spines.values():
                spine.set_visible(False)
        
        for idx, sector in enumerate(sectors[:3]):
            inner_ax = axes[idx]
            sector_data = bh_merged[bh_merged['newta_sector_name'] == sector]
            bands = sector_data['newbh_cell_fdd_band'].unique()
            
            for band_idx, band in enumerate(bands):
                band_data = sector_data[sector_data['newbh_cell_fdd_band'] == band].sort_values('newbh_date')
                if band_data.empty:
                    continue
                band_str = str(int(band)) if pd.notna(band) else "Unknown"
                color = self.color_palette[band_idx % len(self.color_palette)]
                
                inner_ax.plot(band_data['newbh_date'], band_data['cqi'], 
                              marker='o', label=f'L{band_str}', color=color, linewidth=2, markersize=5)
            
            inner_ax.set_title(f'Sector {sector}', fontsize=12, fontweight='bold', pad=10)
            inner_ax.set_xlabel('Date', fontsize=10)
            inner_ax.set_ylabel('Average CQI', fontsize=10)
            inner_ax.legend(fontsize=8, loc='best', frameon=False)
            inner_ax.grid(True, alpha=0.2, linestyle='--')
            inner_ax.tick_params(axis='x', rotation=45, labelsize=9)
            inner_ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            inner_ax.set_ylim(0, 15)
            if num_sectors > 1:
                inner_ax.set_facecolor('#f9f9f9')

    def _plot_spectral_efficiency_on_axis(self, ax, wd_merged):
        wd_merged['spectral_eff'] = np.where(
            wd_merged['newwd_spectral_efficiency_dl_den'] > 0,
            wd_merged['newwd_spectral_efficiency_dl_num'] / wd_merged['newwd_spectral_efficiency_dl_den'],
            np.nan
        )
        
        sectors = wd_merged['newta_sector_name'].dropna().unique()
        if len(sectors) == 0:
            return
        
        num_sectors = min(3, len(sectors))
        axes = []
        
        gap = 0.06
        width = (0.88 - (num_sectors - 1) * gap) / num_sectors if num_sectors > 1 else 0.88
        
        for i in range(num_sectors):
            left = 0.06 + i * (width + gap)
            inner_ax = ax.inset_axes([left, 0.12, width, 0.76])
            inner_ax.set_frame_on(False)
            for spine in inner_ax.spines.values():
                spine.set_visible(False)
            axes.append(inner_ax)
        
        if num_sectors == 1:
            axes = [ax.inset_axes([0.06, 0.12, 0.88, 0.76])]
            axes[0].set_frame_on(False)
            for spine in axes[0].spines.values():
                spine.set_visible(False)
        
        for idx, sector in enumerate(sectors[:3]):
            inner_ax = axes[idx]
            sector_data = wd_merged[wd_merged['newta_sector_name'] == sector]
            bands = sector_data['newwd_cell_fdd_band'].unique()
            
            for band_idx, band in enumerate(bands):
                band_data = sector_data[sector_data['newwd_cell_fdd_band'] == band].sort_values('newwd_date')
                if band_data.empty: continue
                band_str = str(int(band)) if pd.notna(band) else "Unknown"
                color = self.color_palette[band_idx % len(self.color_palette)]
                
                inner_ax.plot(band_data['newwd_date'], band_data['spectral_eff'], 
                              marker='o', label=f'L{band_str}', color=color, linewidth=2, markersize=5)
            
            inner_ax.set_title(f'Sector {sector}', fontsize=12, fontweight='bold', pad=10)
            inner_ax.set_xlabel('Date', fontsize=10)
            inner_ax.set_ylabel('Spectral Efficiency', fontsize=10)
            inner_ax.legend(fontsize=8, loc='best', frameon=False)
            inner_ax.grid(True, alpha=0.2, linestyle='--')
            inner_ax.tick_params(axis='x', rotation=45, labelsize=9)
            inner_ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            inner_ax.set_ylim(bottom=0)
            if num_sectors > 1:
                inner_ax.set_facecolor('#f9f9f9')

    def _plot_qpsk_on_axis(self, ax, bh_merged):
        bh_merged['qpsk'] = np.where(
            bh_merged['newbh_cell_qpsk_rate_den'] > 0,
            (bh_merged['newbh_cell_qpsk_rate_num'] / bh_merged['newbh_cell_qpsk_rate_den']) * 100,
            np.nan
        )
        
        sectors = bh_merged['newta_sector_name'].dropna().unique()
        if len(sectors) == 0:
            return
        
        num_sectors = min(3, len(sectors))
        axes = []
        
        gap = 0.06
        width = (0.88 - (num_sectors - 1) * gap) / num_sectors if num_sectors > 1 else 0.88
        
        for i in range(num_sectors):
            left = 0.06 + i * (width + gap)
            inner_ax = ax.inset_axes([left, 0.12, width, 0.76])
            inner_ax.set_frame_on(False)
            for spine in inner_ax.spines.values():
                spine.set_visible(False)
            axes.append(inner_ax)
        
        if num_sectors == 1:
            axes = [ax.inset_axes([0.06, 0.12, 0.88, 0.76])]
            axes[0].set_frame_on(False)
            for spine in axes[0].spines.values():
                spine.set_visible(False)
        
        for idx, sector in enumerate(sectors[:3]):
            inner_ax = axes[idx]
            sector_data = bh_merged[bh_merged['newta_sector_name'] == sector]
            bands = sector_data['newbh_cell_fdd_band'].unique()
            
            for band_idx, band in enumerate(bands):
                band_data = sector_data[sector_data['newbh_cell_fdd_band'] == band].sort_values('newbh_date')
                if band_data.empty: continue
                band_str = str(int(band)) if pd.notna(band) else "Unknown"
                color = self.color_palette[band_idx % len(self.color_palette)]
                
                inner_ax.plot(band_data['newbh_date'], band_data['qpsk'], 
                              marker='o', label=f'L{band_str}', color=color, linewidth=2, markersize=5)
            
            inner_ax.set_title(f'Sector {sector}', fontsize=12, fontweight='bold', pad=10)
            inner_ax.set_xlabel('Date', fontsize=10)
            inner_ax.set_ylabel('QPSK Rate (%)', fontsize=10)
            inner_ax.legend(fontsize=8, loc='best', frameon=False)
            inner_ax.grid(True, alpha=0.2, linestyle='--')
            inner_ax.tick_params(axis='x', rotation=45, labelsize=9)
            inner_ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            inner_ax.set_ylim(0, 100)
            if num_sectors > 1:
                inner_ax.set_facecolor('#f9f9f9')

    def _plot_timing_advance_on_axis(self, ax, ta_df):
        sectors = ta_df['newta_sector_name'].unique()
        if len(sectors) == 0:
            return
        
        num_sectors = min(3, len(sectors))
        axes = []
        
        gap = 0.06
        width = (0.88 - (num_sectors - 1) * gap) / num_sectors if num_sectors > 1 else 0.88
        
        for i in range(num_sectors):
            left = 0.06 + i * (width + gap)
            inner_ax = ax.inset_axes([left, 0.12, width, 0.76])
            inner_ax.set_frame_on(False)
            for spine in inner_ax.spines.values():
                spine.set_visible(False)
            axes.append(inner_ax)
        
        if num_sectors == 1:
            axes = [ax.inset_axes([0.06, 0.12, 0.88, 0.76])]
            axes[0].set_frame_on(False)
            for spine in axes[0].spines.values():
                spine.set_visible(False)
        
        distance_labels = ["0-78", "78-234", "234-390", "390-546", "546-702", "702-858",
                           "858-1014", "1014-1560", "1560-2106", "2106-2652", "2652-3120",
                           "3120-3900", "3900-6318", "6318-10062", "10062-13962", "13962-20000"]
        
        x_pos = np.arange(len(distance_labels))
        
        for idx, sector in enumerate(sectors[:num_sectors]):
            inner_ax = axes[idx]
            sector_data = ta_df[ta_df['newta_sector_name'] == sector]
            bands = sorted([b for b in sector_data['newta_band'].unique() if pd.notna(b)])
            
            bar_width = 0.8 / len(bands) if bands else 0.8
            
            for band_idx, band in enumerate(bands):
                band_data = sector_data[sector_data['newta_band'] == band].iloc[0]
                band_str = str(int(band))
                color = self.band_colors.get(band_str, "#95A5A6")
                
                distance_cols = [f'newta_{d}_m' for d in ["0_78", "78_234", "234_390", "390_546", "546_702", "702_858",
                                                         "858_1014", "1014_1560", "1560_2106", "2106_2652", "2652_3120",
                                                         "3120_3900", "3900_6318", "6318_10062", "10062_13962", "13962_20000"]]
                values = [band_data[col] if col in band_data and pd.notna(band_data[col]) else 0 for col in distance_cols]
                
                pos = x_pos + band_idx * bar_width - (len(bands)-1)*bar_width/2
                inner_ax.bar(pos, values, width=bar_width, color=color, alpha=0.7, label=f'L{band_str}')
            
            # Twin axis for CDF
            ax2 = inner_ax.twinx()
            for band_idx, band in enumerate(bands):
                band_data = sector_data[sector_data['newta_band'] == band].iloc[0]
                band_str = str(int(band))
                color = self.band_colors.get(band_str, "#95A5A6")
                
                cdf_cols = ['newta_78', 'newta_234', 'newta_390', 'newta_546', 'newta_702', 'newta_858',
                            'newta_1014', 'newta_1560', 'newta_2106', 'newta_2652', 'newta_3120', 'newta_3900',
                            'newta_6318', 'newta_10062', 'newta_13962', 'newta_20000']
                cdf_values = [band_data[col] if col in band_data and pd.notna(band_data[col]) else 0 for col in cdf_cols]
                
                ax2.plot(x_pos, cdf_values, marker='o', color=color, linewidth=2, label=f'L{band_str} CDF')
            
            ax2.axhline(y=90, color='red', linestyle='--', linewidth=1.5, alpha=0.7)
            ax2.set_ylim(0, 105)
            ax2.grid(True, alpha=0.2, linestyle='--')
            
            inner_ax.set_title(f'Sector {sector}', fontsize=12, fontweight='bold', pad=10)
            inner_ax.set_xlabel('Distance (m)', fontsize=10)
            inner_ax.set_ylabel('Number of Samples', fontsize=10)
            inner_ax.set_xticks(x_pos)
            inner_ax.set_xticklabels(distance_labels, rotation=45, ha='right', fontsize=8)
            inner_ax.legend(loc='upper left', fontsize=8, ncol=2, frameon=False)
            ax2.legend(loc='upper right', fontsize=8, frameon=False)
            
            if num_sectors > 1:
                inner_ax.set_facecolor('#f9f9f9')

    def copy_range_with_excel(self, source_file, template_file, output_file):
        """Copy range B6:V34 as picture and paste to both 'SSH Achievement' and 'Justification' sheets"""
        ps_file = None
        try:
            ps_script = f'''
    $ErrorActionPreference = "Stop"

    try {{
        $excel = New-Object -ComObject Excel.Application
        $excel.Visible = $false
        $excel.DisplayAlerts = $false

        Write-Host "Opening source file: {source_file}"
        $sourceWb = $excel.Workbooks.Open("{source_file}")
        $sourceWs = $sourceWb.Worksheets.Item("SSH Achievement")

        Write-Host "Copying range B6:V34 as picture..."
        $range = $sourceWs.Range("B6:V34")
        $range.CopyPicture(1, 2)  # Picture format, as shown on screen

        Write-Host "Opening template: {template_file}"
        $templateWb = $excel.Workbooks.Open("{template_file}")

        # === Paste to SSH Achievement sheet ===
        if ($templateWb.Worksheets | Where-Object {{ $_.Name -eq "SSH Achievement" }}) {{
            $sshWs = $templateWb.Worksheets.Item("SSH Achievement")
            $sshWs.Activate()
            Write-Host "Pasting picture to SSH Achievement at C9..."
            $pasteRangeSSH = $sshWs.Range("C9")
            $pasteRangeSSH.Select()
            $sshWs.Paste()
        }} else {{
            Write-Host "WARNING: Sheet 'SSH Achievement' not found in template"
        }}

        # === Paste to Justification sheet ===
        if ($templateWb.Worksheets | Where-Object {{ $_.Name -eq "Justification" }}) {{
            $justWs = $templateWb.Worksheets.Item("Justification")
            $justWs.Activate()
            Write-Host "Pasting picture to Justification at C9..."
            $pasteRangeJust = $justWs.Range("C9")
            $pasteRangeJust.Select()
            $justWs.Paste()
        }} else {{
            Write-Host "WARNING: Sheet 'Justification' not found in template"
        }}

        Write-Host "Saving as: {output_file}"
        $templateWb.SaveAs("{output_file}", 51)  # xlOpenXMLWorkbook

        Write-Host "Closing workbooks..."
        $templateWb.Close($false)
        $sourceWb.Close($false)
        $excel.Quit()

        # Cleanup COM objects
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($sourceWs) | Out-Null
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($sourceWb) | Out-Null
        if ($sshWs) {{ [System.Runtime.Interopservices.Marshal]::ReleaseComObject($sshWs) | Out-Null }}
        if ($justWs) {{ [System.Runtime.Interopservices.Marshal]::ReleaseComObject($justWs) | Out-Null }}
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($templateWb) | Out-Null
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
        [System.GC]::Collect()
        [System.GC]::WaitForPendingFinalizers()

        Write-Host "SUCCESS: Picture pasted to both sheets"
        exit 0
    }} catch {{
        Write-Host "ERROR: $_"
        exit 1
    }}
    '''

            ps_file = self.output_folder / "temp_copy.ps1"
            with open(ps_file, 'w', encoding='utf-8') as f:
                f.write(ps_script)

            print("  📝 Running PowerShell script to paste picture to both sheets...")
            result = subprocess.run(
                ["powershell", "-ExecutionPolicy", "Bypass", "-File", str(ps_file)],
                capture_output=True,
                text=True,
                encoding='utf-8',
                timeout=120
            )

            print(f"  PowerShell stdout: {result.stdout.strip()}")
            if result.stderr:
                print(f"  PowerShell stderr: {result.stderr.strip()}")
            print(f"  Return code: {result.returncode}")

            output_path = Path(output_file)
            if result.returncode == 0 and output_path.exists():
                print(f"  ✓ Output file created with picture in both sheets: {output_path.name}")
                return True
            else:
                print(f"  ✗ PowerShell failed (code {result.returncode})")
                return False

        except subprocess.TimeoutExpired:
            print("  ✗ PowerShell script timed out")
            return False
        except Exception as e:
            print(f"  ✗ PowerShell execution error: {e}")
            return False
        finally:
            if ps_file and ps_file.exists():
                try:
                    ps_file.unlink()
                except PermissionError:
                    print(f"  ⚠ Could not delete temp script (locked): {ps_file.name}")
                except Exception as e:
                    print(f"  ⚠ Failed to delete temp script: {e}")

    def get_conditional_values(self, ws):
        k_values = ["99.00", "98.00", "1.00", "97.00", "8.00", "50.00", "1.10",
                    "1.00", "120.00", "5.00", "1.00", "20.00", "-105.00", "30.00"]
        n_values = ["99.00", "98.00", "1.00", "97.00", "9.00", "40.00", "1.50",
                    "1.00", "120.00", "10.00", "1.50", "35.00", "-105.00", "30.00",
                    "98.00", "98.00", "5.00"]
        q_values = ["99.00", "98.00", "1.00", "97.00", "9.00", "40.00", "1.70",
                    "1.00", "120.00", "10.00", "1.50", "35.00", "-105.00", "30.00"]
        t_values = ["99.00", "98.00", "1.00", "97.00", "10.00", "40.00", "1.90",
                    "1.00", "120.00", "10.00", "1.50", "35.00", "-105.00", "30.00"]
        
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        try:
            if ws['M28'].value and float(ws['M28'].value) > 0:
                for idx, i in enumerate(range(11, 25)):
                    cell = ws[f'K{i}']
                    cell.value = k_values[idx]
                    cell.fill = white_fill
                    cell.number_format = '@'
            else:
                for i in range(11, 25):
                    cell = ws[f'K{i}']
                    cell.number_format = '@'
                    # cell.fill = white_fill
        except Exception as e:
            print(f"  ⚠ Warning M28: {e}")
        
        # P28 -> N11:N27
        try:
            if ws['P28'].value and float(ws['P28'].value) > 0:
                for idx, i in enumerate(range(11, 28)):
                    cell = ws[f'N{i}']
                    cell.value = n_values[idx]
                    cell.fill = white_fill 
                    cell.number_format = '@'
            else:
                for i in range(11, 28):
                    cell = ws[f'N{i}']
                    # cell.value = "-"
                    cell.number_format = '@'
                    # cell.fill = white_fill
        except Exception as e:
            print(f"  ⚠ Warning P28: {e}")
        
        # S28 -> Q11:Q24
        try:
            if ws['S28'].value and float(ws['S28'].value) > 0:
                for idx, i in enumerate(range(11, 25)):
                    cell = ws[f'Q{i}']
                    cell.value = q_values[idx]
                    cell.fill = white_fill
                    cell.number_format = '@'
            else:
                for i in range(11, 25):
                    cell = ws[f'Q{i}']
                    # cell.value = "-"
                    cell.number_format = '@'
                    # cell.fill = white_fill
        except Exception as e:
            print(f"  ⚠ Warning S28: {e}")
        
        # V28 -> T11:T24
        try:
            if ws['V28'].value and float(ws['V28'].value) > 0:
                for idx, i in enumerate(range(11, 25)):
                    cell = ws[f'T{i}']
                    cell.value = t_values[idx]
                    cell.fill = white_fill
                    cell.number_format = '@'
            else:
                for i in range(11, 25):
                    cell = ws[f'T{i}']
                    # cell.value = "-"
                    cell.number_format = '@'
                    # cell.fill = white_fill
        except Exception as e:
            print(f"  ⚠ Warning V28: {e}")
            
    def list_xlsx_files(self):
        xlsx_files = sorted(self.input_folder.glob("*.xlsx"))
        print(f"Found {len(xlsx_files)} Excel files:")
        for i, file in enumerate(xlsx_files, 1):
            print(f"  {i}. {file.name}")
        return xlsx_files
    
    def extract_cluster_tower(self, text):
        cluster_match = re.search(r'Cluster\s*:\s*([^T]+?)Tower', text)
        tower_match = re.search(r'Tower\s*:\s*([^\s(]+)', text)
        
        cluster = cluster_match.group(1).strip() if cluster_match else "Unknown"
        tower = tower_match.group(1).strip() if tower_match else "Unknown"
        
        return cluster, tower

    def process_file(self, file_path):
        print(f"\nProcessing: {file_path.name}")
        
        try:
            wb = openpyxl.load_workbook(file_path)
            
            if "Cluster & Tower" in wb.sheetnames:
                ws = wb["Cluster & Tower"]
                ws.title = "SSH Achievement"
                print(f"  ✓ Renamed 'Cluster & Tower' to 'SSH Achievement'")
            else:
                print(f"  ✗ Sheet 'Cluster & Tower' not found")
                return None
            
            sheets_to_remove = [s for s in wb.sheetnames if s != "SSH Achievement"]
            for sheet in sheets_to_remove:
                wb.remove(wb[sheet])
            print(f"  ✓ Removed {len(sheets_to_remove)} other sheets")
            
            b6_value = ws['B6'].value or ""
            cluster, tower = self.extract_cluster_tower(str(b6_value))
            print(f"  ✓ Cluster: {cluster} | Tower: {tower}")
            
            self.get_conditional_values(ws)
            
            timestamp = datetime.now().strftime("%H%M%S")
            modified_source = self.output_folder / f"temp_{file_path.stem}_{timestamp}.xlsx"
            wb.save(modified_source)
            wb.close()
            
            time.sleep(1)
            if not modified_source.exists():
                raise FileNotFoundError(f"Temp file failed to save: {modified_source}")
            print(f"  ✓ Saved temp source: {modified_source.name}")

            conn = self.connect_db()
            ta_df = self.query_ta_data(conn, tower)
            wd_df = self.query_wd_data(conn, tower)
            bh_df = self.query_bh_data(conn, tower)
            conn.close()
            
            ta_df, wd_merged, bh_merged = self.merge_data(ta_df, wd_df, bh_df)
            
            if not self.template_path.exists():
                print(f"  ✗ Template not found: {self.template_path}")
                return None
            
            timestamp_full = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"SSH_Achievement_Report_NS_{cluster}_{tower}_{timestamp_full}.xlsx"
            output_path = self.output_folder / output_filename
            
            print("  ⚡ Step 1: Creating base report with Excel automation...")
            success = self.copy_range_with_excel(
                str(modified_source),
                str(self.template_path),
                str(output_path)
            )
            
            if not success or not output_path.exists():
                result = self.create_report_alternative(
                    file_path, cluster, tower, 
                    ta_df, wd_merged, bh_merged, wd_df, output_path
                )
            else:
                self.add_combined_chart_to_report(
                    output_path, ta_df, wd_merged, bh_merged, cluster, tower, wd_df
                )
                result = output_path

            modified_source.unlink(missing_ok=True)
            
            if result and result.exists():
                print(f"  ✓ Successfully created: {result.name}")
                return result
            else:
                print(f"  ✗ Final report not created")
                return None
                
        except Exception as e:
            print(f"  ✗ Error processing {file_path.name}: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def add_combined_chart_to_report(self, report_path, ta_df, wd_merged, bh_merged, cluster, tower, wd_df):
        try:
            wb = openpyxl.load_workbook(report_path)
            
            # Chart tetap di Justification
            if "Justification" in wb.sheetnames:
                just_sheet = wb["Justification"]
                chart_bytes = self.create_combined_chart(ta_df, wd_merged, bh_merged, cluster, tower)
                if chart_bytes:
                    img = XLImage(BytesIO(chart_bytes))
                    img.anchor = 'C73'
                    just_sheet.add_image(img)
                    print("  ✓ Added Combined Chart to Justification!C73")

            # Update cluster & tower di SSH Achievement
            if "SSH Achievement" in wb.sheetnames:
                ssh_sheet = wb["SSH Achievement"]
                ssh_sheet['E34'] = cluster
                ssh_sheet['E34'].font = Font(bold=True, size=12)
                ssh_sheet['E35'] = tower
                ssh_sheet['E35'].font = Font(bold=True, size=12)

            wb.save(report_path)
            wb.close()

            self.add_wd_latest_data_to_report(report_path, wd_df)

        except Exception as e:
            print(f"  ⚠ Could not complete post-processing: {e}")

    def create_report_alternative(self, original_file, cluster, tower, ta_df, wd_merged, bh_merged, wd_df,suggested_path):
        print("  ⚡ Using alternative Python-only method...")
        try:
            template_wb = openpyxl.load_workbook(self.template_path)
            
            # Cari sheet yang benar di template (biasanya "SSH Achievement", tapi cek dulu)
            if "SSH Achievement" in template_wb.sheetnames:
                template_ws = template_wb["SSH Achievement"]
            else:
                # Ambil sheet pertama jika tidak ada
                template_ws = template_wb.active
                print(f"  ⚠ Sheet 'SSH Achievement' not found in template, using '{template_ws.title}'")

            source_wb = openpyxl.load_workbook(original_file)
            source_ws = source_wb["SSH Achievement"]

            # Copy range B6:V34 ke C8:C36 (geser 1 kolom kanan, 2 baris bawah)
            for row_idx, row in enumerate(source_ws['B6:V34'], start=8):  # mulai dari row 8 di target
                for cell in row:
                    target_cell = template_ws.cell(row=row_idx, column=cell.column + 1)  # kolom +1 (B→C)
                    target_cell.value = cell.value
                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.border = copy(cell.border)
                        target_cell.fill = copy(cell.fill)
                        target_cell.number_format = cell.number_format
                        target_cell.protection = copy(cell.protection)
                        target_cell.alignment = copy(cell.alignment)

            source_wb.close()

            # Update cluster & tower
            template_ws['E34'] = cluster
            template_ws['E34'].font = Font(bold=True, size=12)
            template_ws['E35'] = tower
            template_ws['E35'].font = Font(bold=True, size=12)

            # Tambah chart ke sheet Justification
            if "Justification" in template_wb.sheetnames:
                just_sheet = template_wb["Justification"]
                chart_bytes = self.create_combined_chart(ta_df, wd_merged, bh_merged, cluster, tower)
                if chart_bytes:
                    img = XLImage(BytesIO(chart_bytes))
                    img.anchor = 'B46'
                    just_sheet.add_image(img)
                    print("  ✓ Added Combined Chart at B46 (alternative method)")

            # Save dengan nama baru
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt_name = f"SSH_Achievement_Report_NS_{cluster}_{tower}_{timestamp}_ALT.xlsx"
            alt_path = self.output_folder / alt_name
            template_wb.save(alt_path)
            template_wb.close()
            self.add_wd_latest_data_to_report(alt_path, wd_df)
            print(f"  ✓ Created alternative report: {alt_name}")
            return alt_path

        except Exception as e:
            print(f"  ✗ Alternative method failed: {e}")
            import traceback
            traceback.print_exc()
            return None

    def add_wd_latest_data_to_report(self, report_path, wd_df):
        """Add latest unique WD data (max date) to sheet 'SSH Achievement' starting at C53 without header + border"""
        try:
            if wd_df is None or wd_df.empty:
                print("  ℹ No WD data to add")
                return

            # Kolom yang dibutuhkan
            cols_needed = [
                'newwd_date', 'newwd_operator', 'newwd_moentity', 'newwd_enodeb_fdd_msc',
                'newwd_cell_fdd_system', 'newwd_cell_fdd_txrxmode', 'newwd_cell_fdd_vendor', 'newwd_cell_fdd_band'
            ]

            # Pastikan semua kolom ada
            available_cols = [col for col in cols_needed if col in wd_df.columns]
            if len(available_cols) < len(cols_needed):
                print(f"  ⚠ Some WD columns missing: {set(cols_needed) - set(available_cols)}")
            
            wd_filtered = wd_df[available_cols].copy()

            # Konversi tanggal untuk filtering
            if 'newwd_date' in wd_filtered.columns:
                wd_filtered['date_str'] = wd_filtered['newwd_date'].dt.strftime('%Y-%m-%d')
                max_date = wd_filtered['date_str'].max()
                latest_data = wd_filtered[wd_filtered['date_str'] == max_date]
            else:
                latest_data = wd_filtered

            # Hapus duplikat & bersihkan
            latest_data = latest_data.drop_duplicates().drop(columns=['date_str'], errors='ignore')
            latest_data = latest_data[cols_needed]  # urutkan sesuai request

            if latest_data.empty:
                print("  ℹ No unique latest WD data found")
                return

            # Load workbook
            wb = openpyxl.load_workbook(report_path)
            
            if "SSH Achievement" not in wb.sheetnames:
                print("  ⚠ Sheet 'SSH Achievement' not found for WD data")
                wb.close()
                return

            sheet = wb["SSH Achievement"]  # <--- HANYA DI SINI

            # Tulis mulai C53
            start_row = 53
            start_col = 3  # C

            for r_idx, row in enumerate(latest_data.itertuples(index=False, name=None), start=start_row):
                for c_idx, value in enumerate(row, start=start_col):
                    cell = sheet.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    if c_idx == start_col and value is not None:
                        try:
                            cell.number_format = 'YYYY-MM-DD'
                        except:
                            pass

            # Hitung range terisi
            num_rows = len(latest_data)
            end_row = start_row + num_rows - 1
            end_col_letter = openpyxl.utils.get_column_letter(start_col + len(cols_needed) - 1)
            range_str = f"C{start_row}:{end_col_letter}{end_row}"

            # Apply border tipis
            
            thin = Side(style='thin')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for row in sheet[range_str]:
                for cell in row:
                    cell.border = border

            wb.save(report_path)
            wb.close()
            print(f"  ✓ Added {num_rows} unique latest WD records to SSH Achievement!C53:{end_col_letter}{end_row}")

        except Exception as e:
            print(f"  ⚠ Failed to add WD data to SSH Achievement: {e}")
            import traceback
            traceback.print_exc()

    def process_all_files(self):
        files = self.list_xlsx_files()
        if not files:
            print("No Excel files found in the input folder.")
            return
        
        print(f"\n{'='*60}")
        print("Starting batch processing...")
        print(f"{'='*60}")
        
        successful = 0
        failed = 0
        
        for file in files:
            result = self.process_file(file)
            if result:
                successful += 1
            else:
                failed += 1
        
        print(f"\n{'='*60}")
        print("Processing complete!")
        print(f"  ✓ Successful: {successful}")
        print(f"  ✗ Failed: {failed}")
        print(f"  Output folder: {self.output_folder}")
        print(f"{'='*60}")

def main():
    INPUT_FOLDER = "D:\\NEW SITE\\WORK\\REQ\\"
    TEMPLATE_PATH = "./template.xlsx"
    OUTPUT_FOLDER = "./output_reports"
    DB_PATH = "./newdatabase.db"
    
    generator = SSHReportGenerator(INPUT_FOLDER, TEMPLATE_PATH, OUTPUT_FOLDER, DB_PATH)
    generator.process_all_files()

if __name__ == "__main__":
    main()