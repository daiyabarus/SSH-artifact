import os
import re
import sqlite3
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from copy import copy
from io import BytesIO
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
import subprocess
import time
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


class SSHReportGenerator:
    def __init__(self, input_folder, template_path, output_folder, db_path):
        # Resolve all paths to absolute to avoid relative path issues in PowerShell/COM
        self.input_folder = Path(input_folder).resolve()
        self.template_path = Path(template_path).resolve()
        self.output_folder = Path(output_folder).resolve()
        self.db_path = Path(db_path).resolve()
        self.output_folder.mkdir(exist_ok=True, parents=True)

        # Color palettes
        self.color_palette = [
            "#080cec",
            "#ef17e8",
            "#52eb0c",
            "#f59e0b",
            "#10b981",
            "#06b6d4",
            "#ef4444",
            "#a855f7",
            "#14b8a6",
            "#f97316",
        ]

        self.band_colors = {
            "850": "#52eb0c",
            "1800": "#080cec",
            "2100": "#ef17e8",
            "2300F1": "#F39C12",
            "2300F2": "#9B59B6",
        }

        # Mapping nama KPI berdasarkan baris M11:M24
        self.kpi_mapping = {
            11: "Session Setup Success Rate (%)",
            12: "Volte CSSR (%)",
            13: "Erab Drop (%)",
            14: "HOSR (Intra + Inter, LTE to LTE) (%)",
            15: "CQI",
            16: "QPSK Distribution (%)",
            17: "Spectrum Efficiency (bps/Hz)",
            18: "Overall Packet Loss Rate (%)",
            19: "Overall Latency",
            20: "DL Cell Throughput",
            21: "UL Cell Throughput",
            22: "Rank2 (%)",
            23: "UL RSSI (dBm)",
            24: "Last TTI (%)",
        }

        # Mapping kolom ke nama band
        self.band_columns = {
            "M": "LowBand",
            "P": "MidBand_18",
            "S": "MidBand_21",
            "V": "HighBand_23",
        }

        # Mapping band name ke nilai numerik
        self.band_values = {
            "LowBand": [850],
            "MidBand_18": [1800],
            "MidBand_21": [2100],
            "HighBand_23": [2300],
        }

    def connect_db(self):
        if not self.db_path.exists():
            raise FileNotFoundError(f"Database not found: {self.db_path}")
        return sqlite3.connect(self.db_path)

    def parse_date_flexible(self, date_str):
        if pd.isna(date_str):
            return None

        date_formats = [
            "%m/%d/%Y",
            "%Y-%m-%d",
            "%m/%d/%y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%d-%m-%Y",
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(str(date_str).strip(), fmt)
            except ValueError:
                continue
        print(f"  ⚠ Could not parse date: {date_str}")
        return None

    def query_ta_data(self, conn, tower_id):
        query = """
        SELECT * FROM tbl_newta 
        WHERE newta_managed_element = ?
        ORDER BY newta_date DESC
        """
        df = pd.read_sql_query(query, conn, params=(tower_id,))
        return df

    def query_wd_data(self, conn, tower_id):
        query = """
        SELECT newwd_date, 
               newwd_moentity, 
               newwd_cell_fdd_band,
               newwd_enodeb_fdd_msc,
               newwd_operator,
               newwd_cell_fdd_system,
               newwd_cell_fdd_txrxmode,
               newwd_cell_fdd_vendor,
               newwd_spectral_efficiency_dl_num, newwd_spectral_efficiency_dl_den,
               newwd_ul_rssi_num_dbm, newwd_ul_rssi_denom_dbm
        FROM tbl_newwd
        WHERE newwd_enodeb_fdd_msc = ?
        ORDER BY newwd_date
        """
        df = pd.read_sql_query(query, conn, params=(tower_id,))
        df["newwd_date"] = df["newwd_date"].apply(self.parse_date_flexible)
        return df

    def query_bh_data(self, conn, tower_id):
        query = """
        SELECT newbh_date, newbh_moentity, 
               newbh_cell_fdd_band,
               newbh_enodeb_fdd_msc,
               -- Throughput DL/UL
               newbh_pdcp_cell_throughput_dl_num, newbh_pdcp_cell_throughput_dl_denom,
               newbh_pdcp_cell_throughput_ul_num, newbh_pdcp_cell_throughput_ul_den,
               -- CQI & QPSK
               newbh_cell_average_cqi_num, newbh_cell_average_cqi_den,
               newbh_cell_qpsk_rate_num, newbh_cell_qpsk_rate_den,
               -- Rank2
               newbh_cell_mimo_transmission_rank_eq_2_rate_num, newbh_cell_mimo_transmission_rank_eq_2_rate_den,
               -- Last TTI
               newbh_cell_last_tti_ratio_num, newbh_cell_last_tti_ratio_den,
               -- UL RSSI (jika ada di BH, meski biasanya di WD)
               newbh_ul_rssi_num_dbm, newbh_ul_rssi_denom_dbm
        FROM tbl_newbh
        WHERE newbh_enodeb_fdd_msc = ?
        ORDER BY newbh_date
        """
        df = pd.read_sql_query(query, conn, params=(tower_id,))
        df["newbh_date"] = df["newbh_date"].apply(self.parse_date_flexible)
        return df

    def query_kqi_data(self, conn, tower_id):
        query = """
        SELECT newkqi_date, 
               newkqi_tcp_connect_delay_ms, 
               newkqi_tcp_connect_rtt_count_times,
               newkqi_server_side_uplink_tcp_packet_loss_rate,
               newkqi_server_side_downlink_tcp_packet_loss_rate,
               newkqi_client_side_uplink_tcp_packet_loss_rate,
               newkqi_client_side_downlink_tcp_packet_loss_rate
        FROM tbl_newkqi
        WHERE newkqi_swe_l6 = ?
        ORDER BY newkqi_date
        """
        df = pd.read_sql_query(query, conn, params=(tower_id,))
        df["newkqi_date"] = df["newkqi_date"].apply(self.parse_date_flexible)
        return df

    def merge_data(self, ta_df, wd_df, bh_df):
        if ta_df.empty:
            return None, None, None

        wd_merged = None
        if not wd_df.empty:
            wd_merged = pd.merge(
                wd_df,
                ta_df[["newta_eutrancell", "newta_sector", "newta_sector_name"]],
                left_on="newwd_moentity",
                right_on="newta_eutrancell",
                how="left",
            )

        bh_merged = None
        if not bh_df.empty:
            bh_merged = pd.merge(
                bh_df,
                ta_df[["newta_eutrancell", "newta_sector", "newta_sector_name"]],
                left_on="newbh_moentity",
                right_on="newta_eutrancell",
                how="left",
            )

        return ta_df, wd_merged, bh_merged

    def _clean_outer_axis(self, ax):
        """Remove all frames, ticks, and spines from outer axis"""
        ax.set_frame_on(False)
        ax.set_xticks([])
        ax.set_yticks([])
        for spine in ax.spines.values():
            spine.set_visible(False)

    def create_combined_chart(self, ta_df, wd_merged, bh_merged, cluster, tower):
        """Create a single figure with 4 charts arranged in 4 rows"""
        if (
            ta_df.empty
            and (wd_merged is None or wd_merged.empty)
            and (bh_merged is None or bh_merged.empty)
        ):
            return None

        fig = plt.figure(figsize=(22, 28))
        gs = fig.add_gridspec(4, 1, hspace=0.4, wspace=0.1)

        # Row 1: CQI Chart
        ax1 = fig.add_subplot(gs[0])
        self._clean_outer_axis(ax1)
        if bh_merged is not None and not bh_merged.empty:
            self._plot_cqi_on_axis(ax1, bh_merged)
            ax1.set_title("CQI - Busy Hour", fontsize=16, fontweight="bold", pad=20)
        else:
            ax1.text(
                0.5,
                0.5,
                "No CQI Data Available",
                ha="center",
                va="center",
                transform=ax1.transAxes,
                fontsize=12,
                style="italic",
            )
            ax1.set_title("CQI - Busy Hour", fontsize=16, fontweight="bold", pad=20)

        # Row 2: Spectral Efficiency Chart
        ax2 = fig.add_subplot(gs[1])
        self._clean_outer_axis(ax2)
        if wd_merged is not None and not wd_merged.empty:
            self._plot_spectral_efficiency_on_axis(ax2, wd_merged)
            ax2.set_title(
                "Spectral Efficiency - Daily", fontsize=16, fontweight="bold", pad=20
            )
        else:
            ax2.text(
                0.5,
                0.5,
                "No Spectral Efficiency Data Available",
                ha="center",
                va="center",
                transform=ax2.transAxes,
                fontsize=12,
                style="italic",
            )
            ax2.set_title(
                "Spectral Efficiency - Daily", fontsize=16, fontweight="bold", pad=20
            )

        # Row 3: QPSK Chart
        ax3 = fig.add_subplot(gs[2])
        self._clean_outer_axis(ax3)
        if bh_merged is not None and not bh_merged.empty:
            self._plot_qpsk_on_axis(ax3, bh_merged)
            ax3.set_title(
                "QPSK Rate - Busy Hour", fontsize=16, fontweight="bold", pad=20
            )
        else:
            ax3.text(
                0.5,
                0.5,
                "No QPSK Data Available",
                ha="center",
                va="center",
                transform=ax3.transAxes,
                fontsize=12,
                style="italic",
            )
            ax3.set_title(
                "QPSK Rate - Busy Hour", fontsize=16, fontweight="bold", pad=20
            )

        # Row 4: Timing Advance Chart
        ax4 = fig.add_subplot(gs[3])
        self._clean_outer_axis(ax4)
        if not ta_df.empty:
            self._plot_timing_advance_on_axis(ax4, ta_df)
            ax4.set_title("Timing Advance", fontsize=14, fontweight="bold", pad=20)
        else:
            ax4.text(
                0.5,
                0.5,
                "No Timing Advance Data Available",
                ha="center",
                va="center",
                transform=ax4.transAxes,
                fontsize=12,
                style="italic",
            )
            ax4.set_title("Timing Advance", fontsize=14, fontweight="bold", pad=20)

        plt.tight_layout(rect=[0, 0, 1, 1])

        buf = BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()

    def _plot_cqi_on_axis(self, ax, bh_merged):
        bh_merged["cqi"] = np.where(
            bh_merged["newbh_cell_average_cqi_den"] > 0,
            bh_merged["newbh_cell_average_cqi_num"]
            / bh_merged["newbh_cell_average_cqi_den"],
            np.nan,
        )

        sectors = bh_merged["newta_sector_name"].dropna().unique()
        if len(sectors) == 0:
            return

        num_sectors = min(3, len(sectors))
        axes = []

        # Create inner axes directly without intermediate container
        gap = 0.06
        width = (
            (0.88 - (num_sectors - 1) * gap) / num_sectors if num_sectors > 1 else 0.88
        )

        for i in range(num_sectors):
            left = 0.06 + i * (width + gap)
            inner_ax = ax.inset_axes([left, 0.12, width, 0.76])
            inner_ax.set_frame_on(False)
            for spine in inner_ax.spines.values():
                spine.set_visible(False)
            axes.append(inner_ax)

        if num_sectors == 1:
            axes = [ax.inset_axes([0.06, 0.12, 0.88, 0.76])]
            axes[0].set_frame_on(False)
            for spine in axes[0].spines.values():
                spine.set_visible(False)

        for idx, sector in enumerate(sectors[:3]):
            inner_ax = axes[idx]
            sector_data = bh_merged[bh_merged["newta_sector_name"] == sector]
            bands = sector_data["newbh_cell_fdd_band"].unique()

            for band_idx, band in enumerate(bands):
                band_data = sector_data[
                    sector_data["newbh_cell_fdd_band"] == band
                ].sort_values("newbh_date")
                if band_data.empty:
                    continue
                band_str = str(int(band)) if pd.notna(band) else "Unknown"
                color = self.color_palette[band_idx % len(self.color_palette)]

                inner_ax.plot(
                    band_data["newbh_date"],
                    band_data["cqi"],
                    marker="o",
                    label=f"L{band_str}",
                    color=color,
                    linewidth=2,
                    markersize=5,
                )

            inner_ax.set_title(
                f"Sector {sector}", fontsize=12, fontweight="bold", pad=10
            )
            inner_ax.set_xlabel("Date", fontsize=10)
            inner_ax.set_ylabel("Average CQI", fontsize=10)
            inner_ax.legend(fontsize=8, loc="best", frameon=False)
            inner_ax.grid(True, linewidth=1.2, alpha=0.8, linestyle="-", color="gray")
            inner_ax.tick_params(axis="x", rotation=45, labelsize=9)
            inner_ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
            inner_ax.set_ylim(0, 15)
            if num_sectors > 1:
                inner_ax.set_facecolor("#f9f9f9")

    def _plot_spectral_efficiency_on_axis(self, ax, wd_merged):
        wd_merged["spectral_eff"] = np.where(
            wd_merged["newwd_spectral_efficiency_dl_den"] > 0,
            wd_merged["newwd_spectral_efficiency_dl_num"]
            / wd_merged["newwd_spectral_efficiency_dl_den"],
            np.nan,
        )

        sectors = wd_merged["newta_sector_name"].dropna().unique()
        if len(sectors) == 0:
            return

        num_sectors = min(3, len(sectors))
        axes = []

        gap = 0.06
        width = (
            (0.88 - (num_sectors - 1) * gap) / num_sectors if num_sectors > 1 else 0.88
        )

        for i in range(num_sectors):
            left = 0.06 + i * (width + gap)
            inner_ax = ax.inset_axes([left, 0.12, width, 0.76])
            inner_ax.set_frame_on(False)
            for spine in inner_ax.spines.values():
                spine.set_visible(False)
            axes.append(inner_ax)

        if num_sectors == 1:
            axes = [ax.inset_axes([0.06, 0.12, 0.88, 0.76])]
            axes[0].set_frame_on(False)
            for spine in axes[0].spines.values():
                spine.set_visible(False)

        for idx, sector in enumerate(sectors[:3]):
            inner_ax = axes[idx]
            sector_data = wd_merged[wd_merged["newta_sector_name"] == sector]
            bands = sector_data["newwd_cell_fdd_band"].unique()

            for band_idx, band in enumerate(bands):
                band_data = sector_data[
                    sector_data["newwd_cell_fdd_band"] == band
                ].sort_values("newwd_date")
                if band_data.empty:
                    continue
                band_str = str(int(band)) if pd.notna(band) else "Unknown"
                color = self.color_palette[band_idx % len(self.color_palette)]

                inner_ax.plot(
                    band_data["newwd_date"],
                    band_data["spectral_eff"],
                    marker="o",
                    label=f"L{band_str}",
                    color=color,
                    linewidth=2,
                    markersize=5,
                )

            inner_ax.set_title(
                f"Sector {sector}", fontsize=12, fontweight="bold", pad=10
            )
            inner_ax.set_xlabel("Date", fontsize=10)
            inner_ax.set_ylabel("Spectral Efficiency", fontsize=10)
            inner_ax.legend(fontsize=8, loc="best", frameon=False)
            inner_ax.grid(True, linewidth=1.2, alpha=0.8, linestyle="-", color="gray")
            inner_ax.tick_params(axis="x", rotation=45, labelsize=9)
            inner_ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
            inner_ax.set_ylim(bottom=0)
            if num_sectors > 1:
                inner_ax.set_facecolor("#f9f9f9")

    def _plot_qpsk_on_axis(self, ax, bh_merged):
        bh_merged["qpsk"] = np.where(
            bh_merged["newbh_cell_qpsk_rate_den"] > 0,
            (
                bh_merged["newbh_cell_qpsk_rate_num"]
                / bh_merged["newbh_cell_qpsk_rate_den"]
            )
            * 100,
            np.nan,
        )

        sectors = bh_merged["newta_sector_name"].dropna().unique()
        if len(sectors) == 0:
            return

        num_sectors = min(3, len(sectors))
        axes = []

        gap = 0.06
        width = (
            (0.88 - (num_sectors - 1) * gap) / num_sectors if num_sectors > 1 else 0.88
        )

        for i in range(num_sectors):
            left = 0.06 + i * (width + gap)
            inner_ax = ax.inset_axes([left, 0.12, width, 0.76])
            inner_ax.set_frame_on(False)
            for spine in inner_ax.spines.values():
                spine.set_visible(False)
            axes.append(inner_ax)

        if num_sectors == 1:
            axes = [ax.inset_axes([0.06, 0.12, 0.88, 0.76])]
            axes[0].set_frame_on(False)
            for spine in axes[0].spines.values():
                spine.set_visible(False)

        for idx, sector in enumerate(sectors[:3]):
            inner_ax = axes[idx]
            sector_data = bh_merged[bh_merged["newta_sector_name"] == sector]
            bands = sector_data["newbh_cell_fdd_band"].unique()

            for band_idx, band in enumerate(bands):
                band_data = sector_data[
                    sector_data["newbh_cell_fdd_band"] == band
                ].sort_values("newbh_date")
                if band_data.empty:
                    continue
                band_str = str(int(band)) if pd.notna(band) else "Unknown"
                color = self.color_palette[band_idx % len(self.color_palette)]

                inner_ax.plot(
                    band_data["newbh_date"],
                    band_data["qpsk"],
                    marker="o",
                    label=f"L{band_str}",
                    color=color,
                    linewidth=2,
                    markersize=5,
                )

            inner_ax.set_title(
                f"Sector {sector}", fontsize=12, fontweight="bold", pad=10
            )
            inner_ax.set_xlabel("Date", fontsize=10)
            inner_ax.set_ylabel("QPSK Rate (%)", fontsize=10)
            inner_ax.legend(fontsize=8, loc="best", frameon=False)
            inner_ax.grid(True, linewidth=1.2, alpha=0.8, linestyle="-", color="gray")
            inner_ax.tick_params(axis="x", rotation=45, labelsize=9)
            inner_ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
            inner_ax.set_ylim(0, 100)
            if num_sectors > 1:
                inner_ax.set_facecolor("#f9f9f9")

    def _plot_timing_advance_on_axis(self, ax, ta_df):
        sectors = ta_df["newta_sector_name"].unique()
        if len(sectors) == 0:
            return

        num_sectors = min(3, len(sectors))
        axes = []

        gap = 0.06
        width = (
            (0.88 - (num_sectors - 1) * gap) / num_sectors if num_sectors > 1 else 0.88
        )

        for i in range(num_sectors):
            left = 0.06 + i * (width + gap)
            inner_ax = ax.inset_axes([left, 0.12, width, 0.76])
            inner_ax.set_frame_on(False)
            for spine in inner_ax.spines.values():
                spine.set_visible(False)
            axes.append(inner_ax)

        if num_sectors == 1:
            axes = [ax.inset_axes([0.06, 0.12, 0.88, 0.76])]
            axes[0].set_frame_on(False)
            for spine in axes[0].spines.values():
                spine.set_visible(False)

        distance_labels = [
            "0-78",
            "78-234",
            "234-390",
            "390-546",
            "546-702",
            "702-858",
            "858-1014",
            "1014-1560",
            "1560-2106",
            "2106-2652",
            "2652-3120",
            "3120-3900",
            "3900-6318",
            "6318-10062",
            "10062-13962",
            "13962-20000",
        ]

        x_pos = np.arange(len(distance_labels))

        for idx, sector in enumerate(sectors[:num_sectors]):
            inner_ax = axes[idx]
            sector_data = ta_df[ta_df["newta_sector_name"] == sector]
            bands = sorted(
                [b for b in sector_data["newta_band"].unique() if pd.notna(b)]
            )

            bar_width = 0.8 / len(bands) if bands else 0.8

            for band_idx, band in enumerate(bands):
                band_data = sector_data[sector_data["newta_band"] == band].iloc[0]
                band_str = str(int(band))
                color = self.band_colors.get(band_str, "#95A5A6")

                distance_cols = [
                    f"newta_{d}_m"
                    for d in [
                        "0_78",
                        "78_234",
                        "234_390",
                        "390_546",
                        "546_702",
                        "702_858",
                        "858_1014",
                        "1014_1560",
                        "1560_2106",
                        "2106_2652",
                        "2652_3120",
                        "3120_3900",
                        "3900_6318",
                        "6318_10062",
                        "10062_13962",
                        "13962_20000",
                    ]
                ]
                values = [
                    band_data[col]
                    if col in band_data and pd.notna(band_data[col])
                    else 0
                    for col in distance_cols
                ]

                pos = x_pos + band_idx * bar_width - (len(bands) - 1) * bar_width / 2
                inner_ax.bar(
                    pos,
                    values,
                    width=bar_width,
                    color=color,
                    alpha=0.7,
                    label=f"L{band_str}",
                )

            # Twin axis for CDF
            ax2 = inner_ax.twinx()
            for band_idx, band in enumerate(bands):
                band_data = sector_data[sector_data["newta_band"] == band].iloc[0]
                band_str = str(int(band))
                color = self.band_colors.get(band_str, "#95A5A6")

                cdf_cols = [
                    "newta_78",
                    "newta_234",
                    "newta_390",
                    "newta_546",
                    "newta_702",
                    "newta_858",
                    "newta_1014",
                    "newta_1560",
                    "newta_2106",
                    "newta_2652",
                    "newta_3120",
                    "newta_3900",
                    "newta_6318",
                    "newta_10062",
                    "newta_13962",
                    "newta_20000",
                ]
                cdf_values = [
                    band_data[col]
                    if col in band_data and pd.notna(band_data[col])
                    else 0
                    for col in cdf_cols
                ]

                ax2.plot(
                    x_pos,
                    cdf_values,
                    marker="o",
                    color=color,
                    linewidth=2,
                    label=f"L{band_str} CDF",
                )

            ax2.axhline(y=90, color="red", linestyle="--", linewidth=1.5, alpha=0.7)
            ax2.set_ylim(0, 105)
            ax2.grid(True, alpha=0.2, linestyle="--")

            inner_ax.set_title(
                f"Sector {sector}", fontsize=12, fontweight="bold", pad=10
            )
            inner_ax.set_xlabel("Distance (m)", fontsize=10)
            inner_ax.set_ylabel("Number of Samples", fontsize=10)
            inner_ax.set_xticks(x_pos)
            inner_ax.set_xticklabels(
                distance_labels, rotation=45, ha="right", fontsize=8
            )
            inner_ax.legend(loc="upper left", fontsize=8, ncol=2, frameon=False)
            ax2.legend(loc="upper right", fontsize=8, frameon=False)

            if num_sectors > 1:
                inner_ax.set_facecolor("#f9f9f9")

    def copy_range_with_excel(self, source_file, template_file, output_file):
        """Copy range B6:V34 as picture and paste to both 'SSH Achievement' and 'Justification' sheets"""
        ps_file = None
        try:
            ps_script = f'''
    $ErrorActionPreference = "Stop"

    try {{
        $excel = New-Object -ComObject Excel.Application
        $excel.Visible = $false
        $excel.DisplayAlerts = $false

        Write-Host "Opening source file: {source_file}"
        $sourceWb = $excel.Workbooks.Open("{source_file}")
        $sourceWs = $sourceWb.Worksheets.Item("SSH Achievement")

        Write-Host "Copying range B6:V34 as picture..."
        $range = $sourceWs.Range("B6:V34")
        $range.CopyPicture(1, 2)  # Picture format, as shown on screen

        Write-Host "Opening template: {template_file}"
        $templateWb = $excel.Workbooks.Open("{template_file}")

        # === Paste to SSH Achievement sheet ===
        if ($templateWb.Worksheets | Where-Object {{ $_.Name -eq "SSH Achievement" }}) {{
            $sshWs = $templateWb.Worksheets.Item("SSH Achievement")
            $sshWs.Activate()
            Write-Host "Pasting picture to SSH Achievement at C9..."
            $pasteRangeSSH = $sshWs.Range("C9")
            $pasteRangeSSH.Select()
            $sshWs.Paste()
        }} else {{
            Write-Host "WARNING: Sheet 'SSH Achievement' not found in template"
        }}

        # === Paste to Justification sheet ===
        if ($templateWb.Worksheets | Where-Object {{ $_.Name -eq "Justification" }}) {{
            $justWs = $templateWb.Worksheets.Item("Justification")
            $justWs.Activate()
            Write-Host "Pasting picture to Justification at C9..."
            $pasteRangeJust = $justWs.Range("C9")
            $pasteRangeJust.Select()
            $justWs.Paste()
        }} else {{
            Write-Host "WARNING: Sheet 'Justification' not found in template"
        }}

        Write-Host "Saving as: {output_file}"
        $templateWb.SaveAs("{output_file}", 51)  # xlOpenXMLWorkbook

        Write-Host "Closing workbooks..."
        $templateWb.Close($false)
        $sourceWb.Close($false)
        $excel.Quit()

        # Cleanup COM objects
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($sourceWs) | Out-Null
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($sourceWb) | Out-Null
        if ($sshWs) {{ [System.Runtime.Interopservices.Marshal]::ReleaseComObject($sshWs) | Out-Null }}
        if ($justWs) {{ [System.Runtime.Interopservices.Marshal]::ReleaseComObject($justWs) | Out-Null }}
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($templateWb) | Out-Null
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
        [System.GC]::Collect()
        [System.GC]::WaitForPendingFinalizers()

        Write-Host "SUCCESS: Picture pasted to both sheets"
        exit 0
    }} catch {{
        Write-Host "ERROR: $_"
        exit 1
    }}
    '''

            ps_file = self.output_folder / "temp_copy.ps1"
            with open(ps_file, "w", encoding="utf-8") as f:
                f.write(ps_script)

            print("  📝 Running PowerShell script to paste picture to both sheets...")
            result = subprocess.run(
                ["powershell", "-ExecutionPolicy", "Bypass", "-File", str(ps_file)],
                capture_output=True,
                text=True,
                encoding="utf-8",
                timeout=120,
            )

            print(f"  PowerShell stdout: {result.stdout.strip()}")
            if result.stderr:
                print(f"  PowerShell stderr: {result.stderr.strip()}")
            print(f"  Return code: {result.returncode}")

            output_path = Path(output_file)
            if result.returncode == 0 and output_path.exists():
                print(
                    f"  ✓ Output file created with picture in both sheets: {output_path.name}"
                )
                return True
            else:
                print(f"  ✗ PowerShell failed (code {result.returncode})")
                return False

        except subprocess.TimeoutExpired:
            print("  ✗ PowerShell script timed out")
            return False
        except Exception as e:
            print(f"  ✗ PowerShell execution error: {e}")
            return False
        finally:
            if ps_file and ps_file.exists():
                try:
                    ps_file.unlink()
                except PermissionError:
                    print(f"  ⚠ Could not delete temp script (locked): {ps_file.name}")
                except Exception as e:
                    print(f"  ⚠ Failed to delete temp script: {e}")

    def get_conditional_values(self, ws):
        k_values = [
            "99.00",
            "98.00",
            "1.00",
            "97.00",
            "8.00",
            "50.00",
            "1.10",
            "1.00",
            "120.00",
            "5.00",
            "1.00",
            "20.00",
            "-105.00",
            "30.00",
        ]
        n_values = [
            "99.00",
            "98.00",
            "1.00",
            "97.00",
            "9.00",
            "40.00",
            "1.50",
            "1.00",
            "120.00",
            "10.00",
            "1.50",
            "35.00",
            "-105.00",
            "30.00",
            "98.00",
            "98.00",
            "5.00",
        ]
        q_values = [
            "99.00",
            "98.00",
            "1.00",
            "97.00",
            "9.00",
            "40.00",
            "1.70",
            "1.00",
            "120.00",
            "10.00",
            "1.50",
            "35.00",
            "-105.00",
            "30.00",
        ]
        t_values = [
            "99.00",
            "98.00",
            "1.00",
            "97.00",
            "10.00",
            "40.00",
            "1.90",
            "1.00",
            "120.00",
            "10.00",
            "1.50",
            "35.00",
            "-105.00",
            "30.00",
        ]

        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

        try:
            if ws["M28"].value and float(ws["M28"].value) > 0:
                for idx, i in enumerate(range(11, 25)):
                    cell = ws[f"K{i}"]
                    cell.value = k_values[idx]
                    cell.fill = white_fill
                    cell.number_format = "@"
            else:
                for i in range(11, 25):
                    cell = ws[f"K{i}"]
                    cell.number_format = "@"
        except Exception as e:
            print(f"  ⚠ Warning M28: {e}")

        try:
            if ws["P28"].value and float(ws["P28"].value) > 0:
                for idx, i in enumerate(range(11, 28)):
                    cell = ws[f"N{i}"]
                    cell.value = n_values[idx]
                    cell.fill = white_fill
                    cell.number_format = "@"
            else:
                for i in range(11, 28):
                    cell = ws[f"N{i}"]
                    cell.number_format = "@"
        except Exception as e:
            print(f"  ⚠ Warning P28: {e}")

        try:
            if ws["S28"].value and float(ws["S28"].value) > 0:
                for idx, i in enumerate(range(11, 25)):
                    cell = ws[f"Q{i}"]
                    cell.value = q_values[idx]
                    cell.fill = white_fill
                    cell.number_format = "@"
            else:
                for i in range(11, 25):
                    cell = ws[f"Q{i}"]
                    cell.number_format = "@"
        except Exception as e:
            print(f"  ⚠ Warning S28: {e}")

        try:
            if ws["V28"].value and float(ws["V28"].value) > 0:
                for idx, i in enumerate(range(11, 25)):
                    cell = ws[f"T{i}"]
                    cell.value = t_values[idx]
                    cell.fill = white_fill
                    cell.number_format = "@"
            else:
                for i in range(11, 25):
                    cell = ws[f"T{i}"]
                    cell.number_format = "@"
        except Exception as e:
            print(f"  ⚠ Warning V28: {e}")

    def list_xlsx_files(self):
        xlsx_files = sorted(self.input_folder.glob("*.xlsx"))
        print(f"Found {len(xlsx_files)} Excel files:")
        for i, file in enumerate(xlsx_files, 1):
            print(f"  {i}. {file.name}")
        return xlsx_files

    def extract_cluster_tower(self, text):
        cluster_match = re.search(r"Cluster\s*:\s*([^T]+?)Tower", text)
        tower_match = re.search(r"Tower\s*:\s*([^\s(]+)", text)

        cluster = cluster_match.group(1).strip() if cluster_match else "Unknown"
        tower = tower_match.group(1).strip() if tower_match else "Unknown"

        return cluster, tower

    def _is_cell_zero_value(self, cell):
        """
        Check if a cell value is effectively zero.
        Handles various formats: 0, 0.00, 0%, -0.00, etc.
        """
        if cell.value is None:
            return False

        value = cell.value

        if isinstance(value, (int, float)):
            try:
                return abs(float(value)) < 1e-10
            except (ValueError, TypeError):
                return False

        if isinstance(value, str):
            str_val = value.strip()

            if not str_val:
                return False

            simple_zeros = [
                "0",
                "0.0",
                "0.00",
                "0.000",
                "0.0000",
                "-0",
                "-0.0",
                "-0.00",
                "-0.000",
                "0%",
                "0.0%",
                "0.00%",
                "0.000%",
                "-0%",
                "-0.0%",
                "-0.00%",
            ]

            if str_val in simple_zeros:
                return True

            # Coba ekstrak angka dari string
            try:
                # Hapus semua karakter non-digit kecuali titik, minus, dan E (scientific notation)
                import re

                clean_str = re.sub(r"[^\d\.\-\+Ee]", "", str_val)

                # Jika setelah pembersihan tidak ada angka
                if not clean_str:
                    return False

                # Konversi ke float
                float_val = float(clean_str)
                return abs(float_val) < 1e-10

            except (ValueError, TypeError):
                # Cek dengan regex patterns
                patterns = [
                    r"^0+$",  # 0, 00, 000
                    r"^0+\.0+$",  # 0.0, 0.00, 0.000
                    r"^-0+\.0+$",  # -0.0, -0.00
                    r"^0+\.0*[%]?$",  # 0%, 0.0%, 0.00%
                    r"^-0+\.0*[%]?$",  # -0%, -0.0%
                    r"^0+\.0*e[+-]?\d+$",  # 0.0e0, 0.00e-10
                ]

                for pattern in patterns:
                    if re.match(pattern, str_val, re.IGNORECASE):
                        return True

                # Cek jika hanya terdiri dari angka 0
                digits_only = re.sub(r"[^\d]", "", str_val)
                if digits_only and all(c == "0" for c in digits_only):
                    return True

        return False

    def process_file(self, file_path):
        print(f"\nProcessing: {file_path.name}")

        try:
            source_wb = openpyxl.load_workbook(file_path)
            source_sheet = None
            if "Cluster & Tower" in source_wb.sheetnames:
                source_sheet = source_wb["Cluster & Tower"]
                sheet_name = "Cluster & Tower"
            elif "SSH Achievement" in source_wb.sheetnames:
                source_sheet = source_wb["SSH Achievement"]
                sheet_name = "SSH Achievement"
            else:
                print(
                    f"  ✗ Tidak menemukan sheet 'Cluster & Tower' atau 'SSH Achievement'"
                )
                source_wb.close()
                return None

            print(f"  ✓ Membaca dari sheet: {sheet_name}")

            # Ambil cluster & tower dari B6
            b6_value = source_sheet["B6"].value or ""
            cluster, tower = self.extract_cluster_tower(str(b6_value))
            print(f"  ✓ Cluster: {cluster} | Tower: {tower}")

            # === DETEKSI FAILED KPI DARI FILE ASLI ===
            failed_kpis = []
            print("  🔍 Memindai KPI failed dari file asli...")

            # Kolom KPI di file asli: M, P, S, V (baris 11-24)
            for row in range(11, 25):
                for col_letter, band_name in self.band_columns.items():
                    cell = source_sheet[f"{col_letter}{row}"]

                    if cell.value is None:
                        continue

                    # Gunakan fungsi deteksi zero value
                    if self._is_cell_zero_value(cell):
                        failed_kpis.append((row, band_name))
                        print(
                            f"    ⚠ KPI failed: {col_letter}{row} = '{cell.value}' → {band_name}"
                        )

            print(f"  📊 Total KPI failed terdeteksi: {len(failed_kpis)}")

            source_wb.close()

            wb = openpyxl.load_workbook(file_path)

            if "Cluster & Tower" in wb.sheetnames:
                ws = wb["Cluster & Tower"]
                ws.title = "SSH Achievement"
                print(f"  ✓ Renamed 'Cluster & Tower' to 'SSH Achievement'")
            else:
                print(f"  ✗ Sheet 'Cluster & Tower' not found")
                return None

            sheets_to_remove = [s for s in wb.sheetnames if s != "SSH Achievement"]
            for sheet in sheets_to_remove:
                wb.remove(wb[sheet])
            print(f"  ✓ Removed {len(sheets_to_remove)} other sheets")

            self.get_conditional_values(ws)

            timestamp = datetime.now().strftime("%H%M%S")
            modified_source = (
                self.output_folder / f"temp_{file_path.stem}_{timestamp}.xlsx"
            )
            wb.save(modified_source)
            wb.close()

            time.sleep(1)
            if not modified_source.exists():
                raise FileNotFoundError(f"Temp file failed to save: {modified_source}")
            print(f"  ✓ Saved temp source: {modified_source.name}")

            # === STEP 3: Query data dari database ===
            conn = self.connect_db()
            ta_df = self.query_ta_data(conn, tower)
            wd_df = self.query_wd_data(conn, tower)
            bh_df = self.query_bh_data(conn, tower)
            conn.close()

            ta_df, wd_merged, bh_merged = self.merge_data(ta_df, wd_df, bh_df)

            if not self.template_path.exists():
                print(f"  ✗ Template not found: {self.template_path}")
                return None

            timestamp_full = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = (
                f"SSH_Achievement_Report_NS_{cluster}_{tower}_{timestamp_full}.xlsx"
            )
            output_path = self.output_folder / output_filename

            print("  ⚡ Step 1: Creating base report with Excel automation...")
            success = self.copy_range_with_excel(
                str(modified_source), str(self.template_path), str(output_path)
            )

            if not success or not output_path.exists():
                result = self.create_report_alternative(
                    file_path,
                    cluster,
                    tower,
                    ta_df,
                    wd_merged,
                    bh_merged,
                    wd_df,
                    output_path,
                    failed_kpis,
                )
            else:
                self.add_combined_chart_to_report_with_precomputed(
                    output_path,
                    ta_df,
                    wd_merged,
                    bh_merged,
                    cluster,
                    tower,
                    wd_df,
                    failed_kpis,
                )
                result = output_path

            modified_source.unlink(missing_ok=True)

            if result and result.exists():
                print(f"  ✓ Successfully created: {result.name}")
                return result
            else:
                print(f"  ✗ Final report not created")
                return None

        except Exception as e:
            print(f"  ✗ Error processing {file_path.name}: {str(e)}")
            import traceback

            traceback.print_exc()
            return None

    def add_combined_chart_to_report(
        self, report_path, ta_df, wd_merged, bh_merged, cluster, tower, wd_df
    ):
        try:
            wb = openpyxl.load_workbook(report_path)

            conn = self.connect_db()
            kqi_df = self.query_kqi_data(conn, tower)
            conn.close()

            failed_kpis = []
            if "SSH Achievement" in wb.sheetnames:
                ssh_sheet = wb["SSH Achievement"]

                shifted_columns = {
                    "N": "LowBand",  # Original M
                    "Q": "MidBand_18",  # Original P
                    "T": "MidBand_21",  # Original S
                    "W": "HighBand_23",  # Original V
                }

                start_row = 14
                end_row = 27

                print(f"  🔍 Checking KPI cells at range N{start_row}:W{end_row}...")

                for row in range(start_row, end_row + 1):
                    for col_letter, band_name in shifted_columns.items():
                        cell = ssh_sheet[f"{col_letter}{row}"]

                        if cell.value is None:
                            continue

                        is_zero = False
                        cell_value = cell.value

                        try:
                            if isinstance(cell_value, str):
                                clean_str = cell_value.strip()
                                clean_str = (
                                    clean_str.replace("%", "").replace(",", "").strip()
                                )
                                if clean_str:
                                    float_val = float(clean_str)
                                    if abs(float_val) < 1e-10:
                                        is_zero = True

                            elif isinstance(cell_value, (int, float)):
                                if abs(float(cell_value)) < 1e-10:
                                    is_zero = True
                        except (ValueError, TypeError):
                            if isinstance(cell_value, str):
                                str_val = cell_value.strip().lower()
                                zero_patterns = [
                                    "0",
                                    "0.0",
                                    "0.00",
                                    "0.000",
                                    "-0",
                                    "-0.0",
                                    "-0.00",
                                    "0%",
                                    "0.0%",
                                    "0.00%",
                                    "-0%",
                                    "-0.0%",
                                ]
                                if str_val in zero_patterns:
                                    is_zero = True

                        if is_zero:
                            original_row = row - 3
                            failed_kpis.append((original_row, band_name))
                            print(
                                f"  ⚠ Failed KPI detected: {col_letter}{row} = '{cell.value}' → {band_name} (Original row: {original_row})"
                            )

            print(f"  📊 Total failed KPIs detected: {len(failed_kpis)}")

            # Tulis daftar failed ke Justification
            self.write_failed_kpi_to_justification(wb, failed_kpis)

            # Tambahkan chart utama
            if "Justification" in wb.sheetnames:
                just_sheet = wb["Justification"]
                main_chart = self.create_combined_chart(
                    ta_df, wd_merged, bh_merged, cluster, tower
                )
                if main_chart:
                    img = XLImage(BytesIO(main_chart))
                    img.anchor = "C49"
                    just_sheet.add_image(img)
                    print("  ✓ Added main combined chart at C49")

                if failed_kpis:
                    failed_chart = self.create_failed_kpi_charts(
                        ta_df, wd_merged, bh_merged, kqi_df, failed_kpis, cluster, tower
                    )
                    if failed_chart:
                        img_failed = XLImage(BytesIO(failed_chart))
                        img_failed.anchor = "C250"
                        just_sheet.add_image(img_failed)
                        print(f"  ✓ Added {len(failed_kpis)} Failed KPI charts at C250")
                else:
                    print("  ℹ No failed KPIs → no failed charts to add")

            # Update cluster & tower
            if "SSH Achievement" in wb.sheetnames:
                ssh_sheet = wb["SSH Achievement"]
                ssh_sheet["E34"] = cluster
                ssh_sheet["E34"].font = Font(bold=True, size=12)
                ssh_sheet["E35"] = tower
                ssh_sheet["E35"].font = Font(bold=True, size=12)

            wb.save(report_path)
            wb.close()
            self.add_wd_latest_data_to_report(report_path, wd_df)

        except Exception as e:
            print(f"  ⚠ Post-processing error: {e}")
            import traceback

            traceback.print_exc()

    def add_combined_chart_to_report_with_precomputed(
        self,
        report_path,
        ta_df,
        wd_merged,
        bh_merged,
        cluster,
        tower,
        wd_df,
        failed_kpis,
    ):
        """Versi baru yang menggunakan precomputed failed KPIs"""
        try:
            wb = openpyxl.load_workbook(report_path)

            conn = self.connect_db()
            kqi_df = self.query_kqi_data(conn, tower)
            conn.close()

            print(f"  📊 Menggunakan precomputed failed KPIs: {len(failed_kpis)} items")

            self.write_failed_kpi_to_justification(wb, failed_kpis)

            if "Justification" in wb.sheetnames:
                just_sheet = wb["Justification"]
                main_chart = self.create_combined_chart(
                    ta_df, wd_merged, bh_merged, cluster, tower
                )
                if main_chart:
                    img = XLImage(BytesIO(main_chart))
                    img.anchor = "C49"
                    just_sheet.add_image(img)
                    print("  ✓ Added main combined chart at C49")

                if failed_kpis:
                    failed_chart = self.create_failed_kpi_charts(
                        ta_df, wd_merged, bh_merged, kqi_df, failed_kpis, cluster, tower
                    )
                    if failed_chart:
                        img_failed = XLImage(BytesIO(failed_chart))
                        img_failed.anchor = "C250"
                        just_sheet.add_image(img_failed)
                        print(f"  ✓ Added {len(failed_kpis)} Failed KPI charts at C250")
                else:
                    print("  ℹ No failed KPIs → no failed charts to add")

            if "SSH Achievement" in wb.sheetnames:
                ssh_sheet = wb["SSH Achievement"]
                ssh_sheet["E34"] = cluster
                ssh_sheet["E34"].font = Font(bold=True, size=12)
                ssh_sheet["E35"] = tower
                ssh_sheet["E35"].font = Font(bold=True, size=12)

            wb.save(report_path)
            wb.close()
            self.add_wd_latest_data_to_report(report_path, wd_df)

        except Exception as e:
            print(f"  ⚠ Post-processing error: {e}")
            import traceback

            traceback.print_exc()

    def create_failed_kpi_charts(
        self, ta_df, wd_merged, bh_merged, kqi_df, failed_kpis, cluster, tower
    ):
        """
        Generate charts for failed KPIs with sector-based grouping and band highlighting.
        Special handling for Overall Latency & Packet Loss (tower-based).
        """
        if not failed_kpis:
            print("  ℹ No failed KPIs detected → skipping failed KPI charts")
            return None

        print(f"  📊 Generating {len(failed_kpis)} Failed KPI chart(s)...")

        if bh_merged is not None and not bh_merged.empty:
            bh_merged["dl_throughput"] = np.where(
                bh_merged["newbh_pdcp_cell_throughput_dl_denom"] > 0,
                bh_merged["newbh_pdcp_cell_throughput_dl_num"]
                / bh_merged["newbh_pdcp_cell_throughput_dl_denom"],
                np.nan,
            )
            bh_merged["ul_throughput"] = np.where(
                bh_merged["newbh_pdcp_cell_throughput_ul_den"] > 0,
                bh_merged["newbh_pdcp_cell_throughput_ul_num"]
                / bh_merged["newbh_pdcp_cell_throughput_ul_den"],
                np.nan,
            )
            bh_merged["cqi"] = np.where(
                bh_merged["newbh_cell_average_cqi_den"] > 0,
                bh_merged["newbh_cell_average_cqi_num"]
                / bh_merged["newbh_cell_average_cqi_den"],
                np.nan,
            )
            bh_merged["qpsk"] = np.where(
                bh_merged["newbh_cell_qpsk_rate_den"] > 0,
                (
                    bh_merged["newbh_cell_qpsk_rate_num"]
                    / bh_merged["newbh_cell_qpsk_rate_den"]
                )
                * 100,
                np.nan,
            )
            bh_merged["rank2"] = np.where(
                bh_merged["newbh_cell_mimo_transmission_rank_eq_2_rate_den"] > 0,
                (
                    bh_merged["newbh_cell_mimo_transmission_rank_eq_2_rate_num"]
                    / bh_merged["newbh_cell_mimo_transmission_rank_eq_2_rate_den"]
                )
                * 100,
                np.nan,
            )
            bh_merged["last_tti"] = np.where(
                bh_merged["newbh_cell_last_tti_ratio_den"] > 0,
                (
                    bh_merged["newbh_cell_last_tti_ratio_num"]
                    / bh_merged["newbh_cell_last_tti_ratio_den"]
                )
                * 100,
                np.nan,
            )

        if wd_merged is not None and not wd_merged.empty:
            wd_merged["spectral_eff"] = np.where(
                wd_merged["newwd_spectral_efficiency_dl_den"] > 0,
                wd_merged["newwd_spectral_efficiency_dl_num"]
                / wd_merged["newwd_spectral_efficiency_dl_den"],
                np.nan,
            )
            wd_merged["ul_rssi"] = np.where(
                wd_merged["newwd_ul_rssi_denom_dbm"] > 0,
                wd_merged["newwd_ul_rssi_num_dbm"]
                / wd_merged["newwd_ul_rssi_denom_dbm"],
                np.nan,
            )

        if kqi_df is not None and not kqi_df.empty:
            kqi_df["latency"] = np.where(
                kqi_df["newkqi_tcp_connect_rtt_count_times"] > 0,
                kqi_df["newkqi_tcp_connect_delay_ms"]
                / kqi_df["newkqi_tcp_connect_rtt_count_times"],
                np.nan,
            )
            kqi_df["packet_loss"] = (
                (
                    kqi_df["newkqi_server_side_uplink_tcp_packet_loss_rate"]
                    + kqi_df["newkqi_server_side_downlink_tcp_packet_loss_rate"]
                )
                / 2
                * 100
            )

        height_per_chart = 5.5
        fig = plt.figure(figsize=(22, max(6, height_per_chart * len(failed_kpis))))
        gs = fig.add_gridspec(len(failed_kpis), 1, hspace=0.6)

        chart_idx = 0
        plotted_any = False

        kpi_config = {
            15: (
                "cqi",
                bh_merged,
                "newbh_cell_fdd_band",
                "newbh_date",
                "Average CQI",
                "CQI",
                0,
                15,
            ),
            16: (
                "qpsk",
                bh_merged,
                "newbh_cell_fdd_band",
                "newbh_date",
                "QPSK Rate",
                "%",
                0,
                100,
            ),
            17: (
                "spectral_eff",
                wd_merged,
                "newwd_cell_fdd_band",
                "newwd_date",
                "Spectral Efficiency",
                "bps/Hz",
                0,
                None,
            ),
            18: (
                "packet_loss",
                kqi_df,
                None,
                "newkqi_date",
                "Overall Packet Loss Rate",
                "%",
                0,
                None,
            ),
            19: (
                "latency",
                kqi_df,
                None,
                "newkqi_date",
                "Overall Latency",
                "ms",
                0,
                None,
            ),
            20: (
                "dl_throughput",
                bh_merged,
                "newbh_cell_fdd_band",
                "newbh_date",
                "DL Cell Throughput",
                "Mbps",
                0,
                None,
            ),
            21: (
                "ul_throughput",
                bh_merged,
                "newbh_cell_fdd_band",
                "newbh_date",
                "UL Cell Throughput",
                "Mbps",
                0,
                None,
            ),
            22: (
                "rank2",
                bh_merged,
                "newbh_cell_fdd_band",
                "newbh_date",
                "Rank2",
                "%",
                0,
                100,
            ),
            23: (
                "ul_rssi",
                wd_merged,
                "newwd_cell_fdd_band",
                "newwd_date",
                "UL RSSI",
                "dBm",
                None,
                None,
            ),
            24: (
                "last_tti",
                bh_merged,
                "newbh_cell_fdd_band",
                "newbh_date",
                "Last TTI",
                "%",
                0,
                100,
            ),
        }

        for row_num, band_name in failed_kpis:
            ax = fig.add_subplot(gs[chart_idx])
            self._clean_outer_axis(ax)

            config = kpi_config.get(row_num)
            if not config:
                ax.text(
                    0.5,
                    0.5,
                    f"KPI Row {row_num} not supported",
                    ha="center",
                    va="center",
                    transform=ax.transAxes,
                    fontsize=12,
                    color="gray",
                )
                chart_idx += 1
                continue

            (
                col_name,
                source_df,
                band_col,
                date_col,
                title_suffix,
                unit,
                ylim_min,
                ylim_max,
            ) = config

            if row_num in [18, 19]:
                if source_df is None or source_df.empty:
                    ax.text(
                        0.5,
                        0.5,
                        "No data available",
                        ha="center",
                        va="center",
                        transform=ax.transAxes,
                        fontsize=12,
                        color="gray",
                    )
                else:
                    ax.plot(
                        source_df[date_col],
                        source_df[col_name],
                        marker="o",
                        linewidth=2.5,
                        markersize=6,
                        color="#d32f2f",
                        label="Tower-wide",
                    )

                    ax.set_title(
                        f"{self.kpi_mapping.get(row_num)}",
                        fontsize=14,
                        fontweight="bold",
                    )
                    ax.set_ylabel(f"{title_suffix} [{unit}]", fontsize=12)
                    ax.set_xlabel("Date", fontsize=12)
                    ax.grid(True, linewidth=1.2, alpha=0.8, linestyle="-", color="gray")
                    ax.tick_params(axis="x", rotation=45)
                    ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
                    ax.legend(fontsize=10, loc="best", frameon=False)

                    if ylim_min is not None:
                        ax.set_ylim(bottom=ylim_min)
                    if ylim_max is not None:
                        ax.set_ylim(top=ylim_max)

                    plotted_any = True

            else:
                if source_df is None or source_df.empty:
                    ax.text(
                        0.5,
                        0.5,
                        "No data available",
                        ha="center",
                        va="center",
                        transform=ax.transAxes,
                        fontsize=12,
                        color="gray",
                    )
                else:
                    allowed_bands = self.band_values.get(band_name, [])

                    if allowed_bands:
                        # filtered_df = source_df[source_df[band_col].isin(allowed_bands)]
                        filtered_df = source_df.copy()
                    else:
                        filtered_df = source_df

                    if filtered_df.empty:
                        ax.text(
                            0.5,
                            0.5,
                            f"No data for {band_name}",
                            ha="center",
                            va="center",
                            transform=ax.transAxes,
                            fontsize=12,
                            color="gray",
                        )
                    else:
                        # Get sectors
                        sectors = filtered_df["newta_sector_name"].dropna().unique()
                        if len(sectors) == 0:
                            ax.text(
                                0.5,
                                0.5,
                                "No sector data",
                                ha="center",
                                va="center",
                                transform=ax.transAxes,
                                fontsize=12,
                                color="gray",
                            )
                        else:
                            num_sectors = min(3, len(sectors))

                            # Create inner axes layout
                            gap = 0.06
                            width = (
                                (0.88 - (num_sectors - 1) * gap) / num_sectors
                                if num_sectors > 1
                                else 0.88
                            )

                            inner_axes = []
                            for i in range(num_sectors):
                                left = 0.06 + i * (width + gap)
                                inner_ax = ax.inset_axes([left, 0.12, width, 0.76])
                                inner_ax.set_frame_on(False)
                                for spine in inner_ax.spines.values():
                                    spine.set_visible(False)
                                inner_axes.append(inner_ax)

                            # Plot each sector
                            for idx, sector in enumerate(sectors[:num_sectors]):
                                inner_ax = inner_axes[idx]
                                sector_data = filtered_df[
                                    filtered_df["newta_sector_name"] == sector
                                ]
                                bands = sector_data[band_col].unique()

                                for band_idx, band in enumerate(bands):
                                    band_data = sector_data[
                                        sector_data[band_col] == band
                                    ].sort_values(date_col)
                                    if band_data.empty:
                                        continue

                                    band_str = (
                                        str(int(band)) if pd.notna(band) else "Unknown"
                                    )

                                    if int(band) in allowed_bands:
                                        color = "#d32f2f"
                                        linewidth = 4
                                        alpha = 1.0
                                    else:
                                        color = self.color_palette[
                                            band_idx % len(self.color_palette)
                                        ]
                                        linewidth = 2
                                        alpha = 0.6

                                    inner_ax.plot(
                                        band_data[date_col],
                                        band_data[col_name],
                                        marker="o",
                                        label=f"L{band_str}",
                                        color=color,
                                        linewidth=linewidth,
                                        markersize=5,
                                        alpha=alpha,
                                    )

                                inner_ax.set_title(
                                    f"Sector {sector}",
                                    fontsize=12,
                                    fontweight="bold",
                                    pad=10,
                                )
                                inner_ax.set_xlabel("Date", fontsize=10)
                                inner_ax.set_ylabel(
                                    f"{title_suffix} [{unit}]", fontsize=10
                                )
                                inner_ax.legend(fontsize=8, loc="best", frameon=False)
                                inner_ax.grid(
                                    True,
                                    linewidth=1.2,
                                    alpha=0.8,
                                    linestyle="-",
                                    color="gray",
                                )
                                inner_ax.tick_params(axis="x", rotation=45, labelsize=9)
                                inner_ax.xaxis.set_major_formatter(
                                    mdates.DateFormatter("%m/%d")
                                )

                                if ylim_min is not None:
                                    inner_ax.set_ylim(bottom=ylim_min)
                                if ylim_max is not None:
                                    inner_ax.set_ylim(top=ylim_max)

                                if num_sectors > 1:
                                    inner_ax.set_facecolor("#f9f9f9")

                            plotted_any = True

            # Set main title
            kpi_name = self.kpi_mapping.get(row_num, f"Row {row_num}")
            if row_num in [18, 19]:
                ax.set_title(
                    f"{kpi_name} (Tower-based)", fontsize=16, fontweight="bold", pad=20
                )
            else:
                ax.set_title(
                    f"{kpi_name} - {band_name}", fontsize=16, fontweight="bold", pad=20
                )

            chart_idx += 1

        if not plotted_any:
            print("  ⚠ No chart was successfully plotted → skipping image")
            plt.close(fig)
            return None

        plt.tight_layout(rect=[0, 0, 1, 1])
        buf = BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        buf.seek(0)
        print("  ✓ Failed KPI charts generated successfully")
        return buf.getvalue()

    def write_failed_kpi_to_justification(self, wb, failed_kpis):
        if "Justification" not in wb.sheetnames or not failed_kpis:
            return

        just_sheet = wb["Justification"]
        start_row = 247

        just_sheet[f"F{start_row - 1}"].value = "KPI Failed List:"
        just_sheet[f"F{start_row - 1}"].font = Font(bold=True, size=14, color="2A2727")

        for idx, (row_num, band_name) in enumerate(failed_kpis):
            kpi_name = self.kpi_mapping.get(row_num, f"Unknown KPI (Row {row_num})")
            display_name = (
                f"{kpi_name} {band_name}" if band_name != "Overall" else kpi_name
            )
            cell = just_sheet[f"F{start_row + idx}"]
            cell.value = display_name
            cell.font = Font(color="2A2727", bold=True, size=11)

    def create_report_alternative(
        self,
        original_file,
        cluster,
        tower,
        ta_df,
        wd_merged,
        bh_merged,
        wd_df,
        suggested_path,
        failed_kpis=None,
    ):
        print("  ⚡ Using alternative Python-only method...")

        if failed_kpis is None:
            failed_kpis = []

        try:
            template_wb = openpyxl.load_workbook(self.template_path)

            if "SSH Achievement" in template_wb.sheetnames:
                template_ws = template_wb["SSH Achievement"]
            else:
                template_ws = template_wb.active
                print(
                    f"  ⚠ Sheet 'SSH Achievement' not found in template, using '{template_ws.title}'"
                )

            source_wb = openpyxl.load_workbook(original_file)
            source_ws = source_wb["SSH Achievement"]

            # Copy range B6:V34 ke C8:C36 (geser 1 kolom kanan, 2 baris bawah)
            for row_idx, row in enumerate(source_ws["B6:V34"], start=8):
                for cell in row:
                    target_cell = template_ws.cell(row=row_idx, column=cell.column + 1)
                    target_cell.value = cell.value
                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.border = copy(cell.border)
                        target_cell.fill = copy(cell.fill)
                        target_cell.number_format = cell.number_format
                        target_cell.protection = copy(cell.protection)
                        target_cell.alignment = copy(cell.alignment)

            source_wb.close()

            # Update cluster & tower
            template_ws["E34"] = cluster
            template_ws["E34"].font = Font(bold=True, size=12)
            template_ws["E35"] = tower
            template_ws["E35"].font = Font(bold=True, size=12)

            # Tambah chart ke sheet Justification
            if "Justification" in template_wb.sheetnames:
                just_sheet = template_wb["Justification"]

                # Main chart
                chart_bytes = self.create_combined_chart(
                    ta_df, wd_merged, bh_merged, cluster, tower
                )
                if chart_bytes:
                    img = XLImage(BytesIO(chart_bytes))
                    img.anchor = "B46"
                    just_sheet.add_image(img)
                    print("  ✓ Added Combined Chart at B46 (alternative method)")

                # Failed KPI charts jika ada
                if failed_kpis:
                    conn = self.connect_db()
                    kqi_df = self.query_kqi_data(conn, tower)
                    conn.close()

                    failed_chart = self.create_failed_kpi_charts(
                        ta_df, wd_merged, bh_merged, kqi_df, failed_kpis, cluster, tower
                    )
                    if failed_chart:
                        img_failed = XLImage(BytesIO(failed_chart))
                        img_failed.anchor = "C250"
                        just_sheet.add_image(img_failed)
                        print(f"  ✓ Added {len(failed_kpis)} Failed KPI charts at C250")

            # Save dengan nama baru
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt_name = (
                f"SSH_Achievement_Report_NS_{cluster}_{tower}_{timestamp}_ALT.xlsx"
            )
            alt_path = self.output_folder / alt_name
            template_wb.save(alt_path)
            template_wb.close()

            self.add_wd_latest_data_to_report(alt_path, wd_df)
            print(f"  ✓ Created alternative report: {alt_name}")
            return alt_path

        except Exception as e:
            print(f"  ✗ Alternative method failed: {e}")
            import traceback

            traceback.print_exc()
            return None

    def add_wd_latest_data_to_report(self, report_path, wd_df):
        """Add latest unique WD data (max date) to sheet 'SSH Achievement' starting at C53 without header + border"""
        try:
            if wd_df is None or wd_df.empty:
                print("  ℹ No WD data to add")
                return

            # Kolom yang dibutuhkan
            cols_needed = [
                "newwd_date",
                "newwd_operator",
                "newwd_moentity",
                "newwd_enodeb_fdd_msc",
                "newwd_cell_fdd_system",
                "newwd_cell_fdd_txrxmode",
                "newwd_cell_fdd_vendor",
                "newwd_cell_fdd_band",
            ]

            available_cols = [col for col in cols_needed if col in wd_df.columns]
            if len(available_cols) < len(cols_needed):
                print(
                    f"  ⚠ Some WD columns missing: {set(cols_needed) - set(available_cols)}"
                )

            wd_filtered = wd_df[available_cols].copy()

            if "newwd_date" in wd_filtered.columns:
                wd_filtered["date_str"] = wd_filtered["newwd_date"].dt.strftime(
                    "%Y-%m-%d"
                )
                max_date = wd_filtered["date_str"].max()
                latest_data = wd_filtered[wd_filtered["date_str"] == max_date]
            else:
                latest_data = wd_filtered

            latest_data = latest_data.drop_duplicates().drop(
                columns=["date_str"], errors="ignore"
            )
            latest_data = latest_data[cols_needed]

            if latest_data.empty:
                print("  ℹ No unique latest WD data found")
                return

            wb = openpyxl.load_workbook(report_path)

            if "SSH Achievement" not in wb.sheetnames:
                print("  ⚠ Sheet 'SSH Achievement' not found for WD data")
                wb.close()
                return

            sheet = wb["SSH Achievement"]

            start_row = 53
            start_col = 3

            for r_idx, row in enumerate(
                latest_data.itertuples(index=False, name=None), start=start_row
            ):
                for c_idx, value in enumerate(row, start=start_col):
                    cell = sheet.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    if c_idx == start_col and value is not None:
                        try:
                            cell.number_format = "YYYY-MM-DD"
                        except:
                            pass

            num_rows = len(latest_data)
            end_row = start_row + num_rows - 1
            end_col_letter = openpyxl.utils.get_column_letter(
                start_col + len(cols_needed) - 1
            )
            range_str = f"C{start_row}:{end_col_letter}{end_row}"

            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for row in sheet[range_str]:
                for cell in row:
                    cell.border = border

            wb.save(report_path)
            wb.close()
            print(
                f"  ✓ Added {num_rows} unique latest WD records to SSH Achievement!C53:{end_col_letter}{end_row}"
            )

        except Exception as e:
            print(f"  ⚠ Failed to add WD data to SSH Achievement: {e}")
            import traceback

            traceback.print_exc()

    def process_all_files(self):
        files = self.list_xlsx_files()
        if not files:
            print("No Excel files found in the input folder.")
            return

        print(f"\n{'=' * 60}")
        print("Starting batch processing...")
        print(f"{'=' * 60}")

        successful = 0
        failed = 0

        for file in files:
            result = self.process_file(file)
            if result:
                successful += 1
            else:
                failed += 1

        print(f"\n{'=' * 60}")
        print("Processing complete!")
        print(f"  ✓ Successful: {successful}")
        print(f"  ✗ Failed: {failed}")
        print(f"  Output folder: {self.output_folder}")
        print(f"{'=' * 60}")


def main():
    INPUT_FOLDER = "D:\\NEW SITE\\WORK\\REQ\\20260109\\"
    TEMPLATE_PATH = "./template.xlsx"
    OUTPUT_FOLDER = "D:\\NEW SITE\\WORK\\REQ\\20260109\\output_reports\\"
    DB_PATH = "./newdatabase.db"

    generator = SSHReportGenerator(INPUT_FOLDER, TEMPLATE_PATH, OUTPUT_FOLDER, DB_PATH)
    generator.process_all_files()


if __name__ == "__main__":
    main()
